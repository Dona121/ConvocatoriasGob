"""
Convocatorias & Proyectos SDP — Streamlit + Supabase
Schema Django v2.
"""
import io, re, math
from datetime import date, datetime, timedelta
import pandas as pd
import streamlit as st
try:
    import folium
    from streamlit_folium import st_folium
    _FOLIUM_OK = True
except ImportError:
    _FOLIUM_OK = False
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

st.set_page_config(
    page_title="Convocatorias & Proyectos SDP",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif}
.block-container{
    padding-top:4.5rem!important;padding-left:2rem!important;
    padding-right:2rem!important;padding-bottom:2rem!important;max-width:100%!important;}
::-webkit-scrollbar{width:8px;height:8px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:#47b1d5;border-radius:10px}
::-webkit-scrollbar-thumb:hover{background:#1754ab}
/* Sidebar — múltiples selectores para Streamlit 1.35–1.41+ */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"]>div,
section[data-testid="stSidebar"]>div:first-child,
section[data-testid="stSidebar"]>div>div,
[data-testid="stSidebarContent"],
[data-testid="stSidebarUserContent"]{
    background:#041e35!important;border-right:none!important;}
section[data-testid="stSidebar"]{
    box-shadow:4px 0 15px rgba(0,0,0,.15)!important;}
section[data-testid="stSidebar"] label{
    color:#fff!important;font-size:.8rem!important;
    text-transform:uppercase;letter-spacing:.05em;}
section[data-testid="stSidebar"] p{
    color:#fff!important;}
/* Dropdowns del sidebar */
section[data-testid="stSidebar"] [data-baseweb="select"] *,
section[data-testid="stSidebar"] [data-baseweb="select"] div,
section[data-testid="stSidebar"] [data-baseweb="select"] span,
section[data-testid="stSidebar"] [data-baseweb="select"] input{
    color:#ffffff!important;}
section[data-testid="stSidebar"] [data-baseweb="select"]>div{
    background:#0d2d4e!important;border-color:#1754ab66!important;}
/* Tags de selección (chips) */
section[data-testid="stSidebar"] [data-baseweb="tag"]{
    background:#1754ab!important;color:#fff!important;}
/* Dropdown abierto (listbox) */
section[data-testid="stSidebar"] [data-baseweb="popover"] [role="listbox"],
section[data-testid="stSidebar"] [role="listbox"]{
    background:#0d2d4e!important;}
section[data-testid="stSidebar"] [role="option"]{
    background:#0d2d4e!important;color:#fff!important;}
section[data-testid="stSidebar"] [role="option"]:hover,
section[data-testid="stSidebar"] [role="option"][aria-selected="true"]{
    background:#1754ab!important;color:#fff!important;}
section[data-testid="stSidebar"] .stButton>button{
    background:#1754ab!important;color:#fff!important;
    border:none!important;transition:all .3s;border-radius:6px!important;}
section[data-testid="stSidebar"] .stButton>button:hover{
    background:#47b1d5!important;color:#041e35!important;}
.stTabs [data-baseweb="tab-list"]{border-bottom:1px solid #e0e0e0;gap:24px}
.stTabs [data-baseweb="tab"]{
    font-weight:600;font-size:.88rem;border-radius:0;
    padding:10px 4px;background:transparent!important;color:#888;border:none;}
.stTabs [aria-selected="true"]{
    background:transparent!important;color:#003d6c!important;
    border-bottom:3px solid #e68878!important;}
.stDownloadButton>button,.stButton>button[kind="primary"]{
    background:#17743d!important;color:white!important;border:none!important;
    border-radius:8px!important;font-weight:600!important;padding:10px 24px!important;
    transition:all .3s!important;}
.stDownloadButton>button:hover,.stButton>button[kind="primary"]:hover{
    background:#005931!important;transform:translateY(-2px);
    box-shadow:0 4px 10px rgba(0,89,49,.3)!important;}
[data-testid="stDataFrame"]{
    border-radius:8px;overflow:hidden;
    box-shadow:0 2px 8px rgba(0,0,0,.04);border:1px solid #e0e0e0;}
/* chat */
.chat-user{background:#e8f0fe;border-radius:16px 16px 4px 16px;padding:12px 16px;
    max-width:80%;font-size:.88rem;color:#1a1a2e;border:1px solid #c5d5f5;}
.chat-ai{background:#fff;border-radius:16px 16px 16px 4px;padding:14px 18px;
    max-width:85%;font-size:.88rem;color:#1a1a1a;border:1px solid #e0e0e0;
    box-shadow:0 2px 6px rgba(0,0,0,.04);}
.chat-ai-label{font-size:.68rem;color:#1754ab;font-weight:700;
    letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px;}
.chat-scroll{max-height:480px;overflow-y:auto;padding:16px;
    background:#f8f9fb;border:1px solid #e0e0e0;border-radius:10px;margin-bottom:16px;}
</style>""", unsafe_allow_html=True)

# ── Credentials ───────────────────────────────────────────────────────────────
try:
    _URL = st.secrets["supabase"]["url"]
    _KEY = st.secrets["supabase"]["key"]
except Exception:
    _URL = "https://keordvjrhcgvnrrvnfa.supabase.co"
    _KEY = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
            "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imtlb3JlZHZqcmhjZ3ZucnJ2bmZhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI1NzA0MDYsImV4cCI6MjA4ODE0NjQwNn0."
            "h9QNpcbiMXZfeheOAVHtYnC4-n8luCg92s-Xd_BFrZA")

try:
    GEMINI_API_KEY = st.secrets["gemini"]["api_key"]
except Exception:
    GEMINI_API_KEY = ""  # configura en Streamlit Cloud → Settings → Secrets
GEMINI_MODEL = "gemini-2.0-flash"

BRAND_COLORS = [
    "#17743d","#1754ab","#cf7000","#47b1d5","#d88c16",
    "#005931","#e68878","#003d6c","#d37e00","#9b5b1e",
]

# ── Supabase ──────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def _sb():
    from supabase import create_client
    return create_client(_URL, _KEY)

def _fetch(table, select="*"):
    client = _sb()
    rows, start = [], 0
    while True:
        resp = client.table(table).select(select).range(start, start+999).execute()
        rows.extend(resp.data)
        if len(resp.data) < 1000: break
        start += 1000
    return rows

def _fdate(val):
    if not val: return "—"
    try: return pd.to_datetime(val).strftime("%d/%m/%Y")
    except: return str(val)

def fmt_money(val):
    try:
        v = float(val)
        if v >= 1e12: return f"${v/1e12:.1f}T"
        if v >= 1e9:  return f"${v/1e9:.1f}B"
        if v >= 1e6:  return f"${v/1e6:.1f}M"
        if v >= 1e3:  return f"${v/1e3:.0f}K"
        return f"${v:,.0f}"
    except: return str(val)

# ── UI helpers ────────────────────────────────────────────────────────────────
def _card(content, title=None):
    hdr = (f'<div style="font-family:\'DM Serif Display\',serif;font-size:1.1rem;color:#003d6c;'
           f'margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #e0e0e0">{title}</div>'
           if title else "")
    return (f'<div style="background:#fff;border:1px solid #e0e0e0;border-radius:10px;'
            f'padding:22px 24px;box-shadow:0 2px 8px rgba(0,0,0,.02)">{hdr}{content}</div>')

def empty_state(texto):
    return (f'<div style="padding:30px 20px;text-align:center;color:#003d6c;'
            f'background:#f0f8fb;border:1px dashed #47b1d5;border-radius:8px;margin:10px 0">'
            f'{texto}</div>')

def bar_chart(data, title, max_bars=20, fmt_val=None):
    data = data.dropna().sort_values(ascending=False).head(max_bars)
    if data.empty: return ""
    mx = data.max() or 1
    rows = ""
    for i, (label, val) in enumerate(data.items()):
        pct   = round(val/mx*100, 1)
        color = BRAND_COLORS[i % len(BRAND_COLORS)]
        disp  = fmt_val(val) if fmt_val else (f"{int(val):,}" if float(val)==int(float(val)) else f"{val:,.1f}")
        rows += (f'<div style="display:flex;align-items:center;margin-bottom:10px;gap:12px">'
                 f'<div style="width:175px;font-size:.75rem;color:#444;font-weight:500;text-align:right;'
                 f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex-shrink:0"title="{label}">{label}</div>'
                 f'<div style="flex:1;background:#f5f5f5;border-radius:4px;height:24px;position:relative">'
                 f'<div style="width:{pct}%;background:{color};height:100%;border-radius:4px"></div>'
                 f'<span style="position:absolute;right:8px;top:4px;font-size:.72rem;font-weight:700;color:#333">{disp}</span>'
                 f'</div></div>')
    return _card(rows, title)

def donut_chart(data, title, top_n=8):
    data  = data.dropna()
    total = data.sum()
    if total == 0: return ""
    top   = data.sort_values(ascending=False).head(top_n)
    cx = cy = 68; r = 52; ir = 28; angle = -90.0; paths = ""
    for i, (_, val) in enumerate(top.items()):
        sw = (val/total)*360; end = angle+sw
        a1, a2 = math.radians(angle), math.radians(end)
        x1,y1   = cx+r*math.cos(a1),  cy+r*math.sin(a1)
        x2,y2   = cx+r*math.cos(a2),  cy+r*math.sin(a2)
        ix1,iy1 = cx+ir*math.cos(a2), cy+ir*math.sin(a2)
        ix2,iy2 = cx+ir*math.cos(a1), cy+ir*math.sin(a1)
        lg = 1 if sw > 180 else 0
        c  = BRAND_COLORS[i % len(BRAND_COLORS)]
        paths += (f'<path d="M{x1:.1f},{y1:.1f} A{r},{r} 0 {lg},1 {x2:.1f},{y2:.1f} '
                  f'L{ix1:.1f},{iy1:.1f} A{ir},{ir} 0 {lg},0 {ix2:.1f},{iy2:.1f} Z" '
                  f'fill="{c}" stroke="#fff" stroke-width="2"/>')
        angle = end
    legend = ""
    for i, (label, val) in enumerate(top.items()):
        pct = round(val/total*100, 1)
        legend += (f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">'
                   f'<div style="width:12px;height:12px;border-radius:50%;'
                   f'background:{BRAND_COLORS[i%len(BRAND_COLORS)]};flex-shrink:0"></div>'
                   f'<div style="font-size:.75rem;color:#555;flex:1;white-space:nowrap;overflow:hidden;'
                   f'text-overflow:ellipsis"title="{label}">{label}</div>'
                   f'<div style="font-size:.75rem;font-weight:700;color:#003d6c">{pct}%</div></div>')
    svg = (f'<svg width="100%" style="max-width:140px;height:auto" viewBox="0 0 136 136">{paths}'
           f'<text x="{cx}" y="{cy+5}" text-anchor="middle" font-size="18" '
           f'font-family="DM Serif Display" fill="#003d6c" font-weight="bold">{int(total)}</text>'
           f'<text x="{cx}" y="{cy+18}" text-anchor="middle" font-size="9" '
           f'font-family="DM Sans" fill="#888">total</text></svg>')
    inner = (f'<div style="display:flex;gap:20px;align-items:center">'
             f'<div style="flex-shrink:0">{svg}</div>'
             f'<div style="flex:1;overflow:hidden">{legend}</div></div>')
    return _card(inner, title)

def kpi(label, value, sub="", style="white", border_color="#47b1d5", flex="1"):
    if style == "dark-blue":
        bg,tc,lc,sc,bs = "#003d6c","#fff","#47b1d5","#a5d6a7","border:none;"
    elif style == "dark-green":
        bg,tc,lc,sc,bs = "#005931","#fff","#7aeb87","#a5d6a7","border:none;"
    else:
        bg,tc,lc,sc = "#fff","#003d6c","#1754ab","#777"
        bs = f"border:1px solid #e0e0e0;border-left:5px solid {border_color};"
    return (f'<div style="flex:{flex};min-width:130px;background:{bg};{bs}'
            f'border-radius:10px;padding:18px 16px;box-shadow:0 3px 8px rgba(0,0,0,.04);'
            f'display:flex;flex-direction:column;justify-content:center;">'
            f'<div style="font-size:.65rem;letter-spacing:.08em;text-transform:uppercase;'
            f'color:{lc};font-weight:700;margin-bottom:4px">{label}</div>'
            f'<div style="font-family:\'DM Serif Display\',serif;font-size:2.1rem;'
            f'color:{tc};line-height:1.1">{value}</div>'
            f'<div style="font-size:.7rem;color:{sc};margin-top:6px">{sub}</div></div>')

def sec_title(text, sub=""):
    s = (f'<div style="font-family:\'DM Serif Display\',serif;font-size:1.45rem;color:#003d6c;'
         f'margin:32px 0 8px;padding-bottom:10px;border-bottom:2px solid #17743d">{text}</div>')
    if sub: s += f'<div style="font-size:.85rem;color:#666;margin-bottom:16px">{sub}</div>'
    return s

def badge(text, color="#1754ab"):
    return (f'<span style="display:inline-block;background:{color}18;color:{color};'
            f'border:1px solid {color}44;border-radius:20px;padding:2px 10px;'
            f'font-size:.72rem;font-weight:600;margin:2px 3px 2px 0">{text}</span>')

def field_row(label, value):
    v = str(value).strip()
    if not v or v in ("", "—", "0", "None"): return ""
    return (f'<div style="display:flex;gap:12px;padding:9px 0;border-bottom:1px solid #f0f0f0">'
            f'<div style="width:170px;flex-shrink:0;font-size:.78rem;font-weight:600;color:#555">{label}</div>'
            f'<div style="flex:1;font-size:.82rem;color:#222">{v}</div></div>')

def stat_grid(*items):
    """items: list of (label, value, color)"""
    cells = ""
    for label, value, color in items:
        cells += (f'<div>'
                  f'<div style="font-size:.66rem;color:#888;text-transform:uppercase;'
                  f'letter-spacing:.06em;margin-bottom:3px">{label}</div>'
                  f'<div style="font-size:1.4rem;font-family:\'DM Serif Display\',serif;'
                  f'color:{color}">{value}</div>'
                  f'</div>')
    return (f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));'
            f'gap:18px;margin-bottom:18px">{cells}</div>')

def _ind_table(inds_list):
    """Render a st.dataframe for a list of indicator dicts."""
    if not inds_list: return
    df_ip = pd.DataFrame(inds_list)
    ic = ["codigo","nombre","vigencia","meta_proyecto","meta_cuatrienio","m2024","m2025","m2026","m2027","responsable_mga"]
    ic = [c for c in ic if c in df_ip.columns]
    st.dataframe(df_ip[ic].rename(columns={
        "codigo":"Código","nombre":"Indicador","vigencia":"Vigencia",
        "meta_proyecto":"Meta proy.","meta_cuatrienio":"Meta cuatrienio",
        "m2024":"2024","m2025":"2025","m2026":"2026","m2027":"2027",
        "responsable_mga":"Responsable MGA",
    }).reset_index(drop=True),
    use_container_width=True, height=min(300, 60+len(inds_list)*42), hide_index=True,
    column_config={"Indicador": st.column_config.TextColumn(width=280)})

# ══════════════════════════════════════════════════════════════════════════════
# LOAD ALL DATA
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=300, show_spinner=False)
def load_all():
    estados    = {r["id"]: r["estado"]            for r in _fetch("contenido_estado")}
    deps       = {r["id"]: r["dependencia"]       for r in _fetch("contenido_dependencia")}
    resps      = {r["id"]: r["responsable"]       for r in _fetch("contenido_responsable")}
    sectores   = {r["id"]: r["sector"]            for r in _fetch("contenido_sectores")}
    segmentos  = {r["id"]: r["segmento"]          for r in _fetch("contenido_segmentos")}
    ubicacs    = {r["id"]: r["ubicacion"]         for r in _fetch("contenido_ubicacion")}
    municipios = {r["id"]: r["municipio"]         for r in _fetch("contenido_municipios")}
    clf_ben    = {r["id"]: r["tipo_beneficiario"] for r in _fetch("contenido_clasificacionbeneficiario")}
    vigencias  = {r["id"]: r["vigencia"]          for r in _fetch("contenido_clasificacionvigencia")}
    clf_aliados= {r["id"]: r["clasificacion_aliado"] for r in _fetch("contenido_clasificacionaliados")}
    aliados_map= {r["id"]: f"{r['aliado']} ({clf_aliados.get(r.get('clasificacion_id'),'—')})"
                  for r in _fetch("contenido_aliados")}
    clf_ind = {r["id"]: {
        "codigo": r["codigo_indicador"], "nombre": r["nombre_indicador"],
        "meta_c": float(r["meta_cuatrienio"] or 0),
        "m2024": r.get("meta_fisica_esperada_2024"), "m2025": r.get("meta_fisica_esperada_2025"),
        "m2026": r.get("meta_fisica_esperada_2026"), "m2027": r.get("meta_fisica_esperada_2027"),
        "resp":  r.get("responsable",""),
    } for r in _fetch("contenido_clasificacionindicadormga")}

    def m2m(table, fk, vk, vmap):
        d = {}
        for r in _fetch(table):
            d.setdefault(r[fk], []).append(vmap.get(r[vk], str(r[vk])))
        return d

    conv_sec = m2m("contenido_convocatorias_sectores",   "convocatorias_id","sectores_id",   sectores)
    conv_seg = m2m("contenido_convocatorias_segmento",   "convocatorias_id","segmentos_id",  segmentos)
    conv_ubi = m2m("contenido_convocatorias_ubicacion",  "convocatorias_id","ubicacion_id",  ubicacs)
    conv_dep = m2m("contenido_convocatorias_dependencia","convocatorias_id","dependencia_id",deps)
    proy_mun = m2m("contenido_proyecto_municipios",      "proyecto_id",     "municipios_id", municipios)

    conv_ali: dict = {}
    for r in _fetch("contenido_convocatorias_aliados"):
        conv_ali.setdefault(r["convocatorias_id"],[]).append(aliados_map.get(r["aliados_id"],"—"))

    ben_d: dict = {}
    for r in _fetch("contenido_beneficiarios"):
        pid = r.get("proyecto_id")
        if pid:
            ben_d.setdefault(pid,[]).append({
                "tipo": clf_ben.get(r.get("beneficiario_id"),"?"),
                "n":    r.get("numero_beneficiarios",0),
            })

    ind_d: dict = {}
    for r in _fetch("contenido_indicadormga"):
        pid = r.get("proyecto_id")
        if pid:
            clf = clf_ind.get(r.get("indicadores_id"),{})
            ind_d.setdefault(pid,[]).append({
                "codigo":          clf.get("codigo",""),
                "nombre":          clf.get("nombre",""),
                "vigencia":        vigencias.get(r.get("vigencia_id"),""),
                "meta_proyecto":   float(r.get("meta_proyecto") or 0),
                "meta_cuatrienio": clf.get("meta_c",0),
                "m2024": clf.get("m2024"), "m2025": clf.get("m2025"),
                "m2026": clf.get("m2026"), "m2027": clf.get("m2027"),
                "responsable_mga": clf.get("resp",""),
            })

    conv_rows = _fetch("contenido_convocatorias")
    conv_list = []
    for r in conv_rows:
        cid = r["id"]
        conv_list.append({
            "id":                 cid,
            "Convocatoria":       r["nombre_convocatoria"],
            "Estado":             estados.get(r.get("estado_id"),"—"),
            "Fecha apertura":     _fdate(r.get("fecha_apertura")),
            "Fecha cierre":       _fdate(r.get("fecha_cierre")),
            "Monto":              float(r.get("monto") or 0),
            "Contacto":           r.get("contacto",""),
            "Qué ofrece":         r.get("que_ofrece","") or "",
            "Quiénes participan": r.get("quienes_pueden_participar","") or "",
            "Público priorizado": r.get("publico_priorizado","") or "",
            "Sectores":           " · ".join(conv_sec.get(cid,[])),
            "Segmentos":          " · ".join(conv_seg.get(cid,[])),
            "Ubicación":          " · ".join(conv_ubi.get(cid,[])),
            "Dependencias":       " · ".join(conv_dep.get(cid,[])),
            "Aliados":            " · ".join(conv_ali.get(cid,[])),
            "N° proyectos":       0,
        })
    df_conv = pd.DataFrame(conv_list) if conv_list else pd.DataFrame()

    proy_rows = _fetch("contenido_proyecto")
    proy_list = []
    for r in proy_rows:
        pid  = r["id"]
        bens = ben_d.get(pid,[])
        inds = ind_d.get(pid,[])
        proy_list.append({
            "id":                  pid,
            "convocatoria_id":     r.get("convocatoria_id"),
            "Proyecto":            r["nombre_proyecto"],
            "BPIN":                r.get("bpin",""),
            "Valor":               float(r.get("valor_proyecto") or 0),
            "Contrapartida":       float(r.get("monto_contrapartida") or 0),
            "Dependencia":         deps.get(r.get("dependencia_id"),"—"),
            "Responsable":         resps.get(r.get("responsable_id"),"—"),
            "Municipios":          " · ".join(proy_mun.get(pid,[])),
            "Total beneficiarios": sum(b["n"] for b in bens),
            "Tipos beneficiarios": ", ".join(f"{b['tipo']} ({b['n']})" for b in bens),
            "N° indicadores MGA":  len(inds),
            "Indicadores MGA":     "; ".join(f"{i['codigo']} – {i['nombre']}" for i in inds),
        })
    df_proy = pd.DataFrame(proy_list) if proy_list else pd.DataFrame()

    if not df_conv.empty and not df_proy.empty and "convocatoria_id" in df_proy.columns:
        cnt = df_proy.groupby("convocatoria_id")["id"].count().to_dict()
        df_conv["N° proyectos"] = df_conv["id"].map(cnt).fillna(0).astype(int)
        val_cnt = df_proy.groupby("convocatoria_id")["Valor"].sum().to_dict()
        df_conv["Valor proyectos"] = df_conv["id"].map(val_cnt).fillna(0)
        df_conv["Cobertura (%)"] = df_conv.apply(
            lambda row: round(row["Valor proyectos"]/row["Monto"]*100,1) if row["Monto"] else None, axis=1)

    proy_names = {r["id"]: r["nombre_proyecto"] for r in proy_rows}
    ind_rows   = []
    for pid, inds in ind_d.items():
        for i in inds:
            ind_rows.append({"Proyecto": proy_names.get(pid,"—"), "proyecto_id": pid, **i})
    df_ind = pd.DataFrame(ind_rows) if ind_rows else pd.DataFrame()

    if not df_proy.empty and not df_conv.empty and "convocatoria_id" in df_proy.columns:
        conv_cols = ["id","Convocatoria","Estado","Monto","Sectores","Segmentos","Ubicación",
                     "Dependencias","Aliados","Fecha apertura","Fecha cierre",
                     "Qué ofrece","Quiénes participan","Público priorizado","Contacto"]
        df_rel = df_proy.merge(
            df_conv[[c for c in conv_cols if c in df_conv.columns]],
            left_on="convocatoria_id", right_on="id", how="left", suffixes=("","_conv")
        ).rename(columns={"Monto":"Monto convocatoria","Estado":"Estado convocatoria"})
        df_rel = df_rel.drop(columns=[c for c in df_rel.columns if c.endswith("_conv")], errors="ignore")
        df_rel["Cobertura (%)"] = df_rel.apply(
            lambda row: round(row["Valor"]/row["Monto convocatoria"]*100,1)
            if row.get("Monto convocatoria",0) else None, axis=1)
    else:
        df_rel = pd.DataFrame()

    return df_conv, df_proy, df_rel, df_ind, ind_d

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(
        '<div style="padding:16px 0 20px">'
        '<div style="font-family:\'DM Serif Display\',serif;font-size:1.8rem;color:#fff;line-height:1.1">SDP</div>'
        '<div style="color:#47b1d5;font-size:.85rem;font-weight:400;margin-top:6px">Convocatorias & Proyectos</div></div>'
        '<hr style="border-color:#1754ab;opacity:.3;margin-bottom:18px">',
        unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LOAD
# ══════════════════════════════════════════════════════════════════════════════
with st.spinner("Conectando con Supabase..."):
    try:
        df_conv, df_proy, df_rel, df_ind, _ind_d = load_all()
    except Exception as e:
        st.error(f"Error al cargar datos: {e}")
        st.stop()

if df_conv.empty and df_proy.empty:
    st.markdown(empty_state("No se encontraron datos en Supabase."), unsafe_allow_html=True)
    st.stop()

# ── Filters ────────────────────────────────────────────────────────────────────
estados_opts  = sorted(df_conv["Estado"].dropna().unique())   if not df_conv.empty else []
sectores_opts = sorted({s.strip() for row in df_conv["Sectores"] if row
                        for s in row.split(" · ") if s.strip()}) if not df_conv.empty else []
dep_opts      = sorted(df_proy["Dependencia"].dropna().unique()) if not df_proy.empty else []

with st.sidebar:
    st.markdown('<div style="font-size:.7rem;letter-spacing:.12em;text-transform:uppercase;'
                'color:#a5d6a7;font-weight:700;margin-bottom:10px">Filtros</div>',
                unsafe_allow_html=True)
    sel_est = st.multiselect("Estado convocatoria", estados_opts,  placeholder="Todos")
    sel_sec = st.multiselect("Sector",              sectores_opts, placeholder="Todos")
    sel_dep = st.multiselect("Dependencia",         dep_opts,      placeholder="Todas")
    st.markdown('<hr style="border-color:#1754ab;opacity:.3;margin:20px 0 16px">', unsafe_allow_html=True)
    if st.button("Refrescar", use_container_width=True):
        st.cache_data.clear(); st.rerun()

df_c = df_conv.copy(); df_p = df_proy.copy()
# Tipos numéricos consistentes para evitar fallos en isin/comparaciones float vs int
for _df, _cols in [(df_c, ["id"]), (df_p, ["id","convocatoria_id"])]:
    for _col in _cols:
        if _col in _df.columns:
            _df[_col] = pd.to_numeric(_df[_col], errors="coerce")
if sel_est:  df_c = df_c[df_c["Estado"].isin(sel_est)]
if sel_sec:  df_c = df_c[df_c["Sectores"].apply(lambda s: any(x in s for x in sel_sec))]
if sel_dep:  df_p = df_p[df_p["Dependencia"].isin(sel_dep)]
if sel_est or sel_sec: df_p = df_p[df_p["convocatoria_id"].isin(df_c["id"])]
if sel_dep:  df_c = df_c[df_c["id"].isin(df_p["convocatoria_id"])]
df_r = df_rel.copy()
if not df_r.empty:
    if sel_est: df_r = df_r[df_r["Estado convocatoria"].isin(sel_est)]
    if sel_dep: df_r = df_r[df_r["Dependencia"].isin(sel_dep)]
df_i = df_ind.copy()
if not df_i.empty and sel_dep:
    df_i = df_i[df_i["proyecto_id"].isin(df_p["id"])]

# ── KPIs globales ──────────────────────────────────────────────────────────────
# ── Notificaciones: convocatorias próximas a cerrar ───────────────────────────
if not df_c.empty and "Fecha cierre" in df_c.columns:
    _hoy = date.today()
    _proximas = []
    for _, _rc in df_c.iterrows():
        _fc_str = str(_rc.get("Fecha cierre","")).strip()
        if not _fc_str or _fc_str == "—": continue
        try:
            _fc = datetime.strptime(_fc_str, "%d/%m/%Y").date()
            _dias = (_fc - _hoy).days
            if 0 <= _dias <= 15:
                _proximas.append((_dias, _rc["Convocatoria"], _fc_str, _rc.get("Estado","—")))
        except Exception:
            pass
    if _proximas:
        _proximas.sort()
        _urgente = any(d <= 5 for d,*_ in _proximas)
        _bg  = "#fff3cd" if not _urgente else "#fde8e8"
        _bdr = "#cf7000" if not _urgente else "#c0392b"
        _ico = "Aviso" if not _urgente else "Urgente"
        _items = "".join(
            f'<div style="display:flex;align-items:center;gap:10px;padding:5px 0;'
            f'border-bottom:1px solid {_bdr}22">'
            f'<span style="background:{_bdr};color:#fff;border-radius:20px;'
            f'padding:1px 9px;font-size:.72rem;font-weight:700;white-space:nowrap">'
            f'{"HOY" if d==0 else f"{d}d"}</span>'
            f'<span style="font-size:.83rem;color:#222;flex:1">{nm}</span>'
            f'<span style="font-size:.75rem;color:#666;white-space:nowrap">cierra {fc}</span>'
            f'</div>'
            for d,nm,fc,_ in _proximas
        )
        st.markdown(
            f'<div style="background:{_bg};border:1px solid {_bdr};border-left:5px solid {_bdr};'
            f'border-radius:8px;padding:14px 20px;margin-bottom:18px">'
            f'<div style="font-weight:700;font-size:.85rem;color:{_bdr};margin-bottom:8px">'
            f'{_ico} — {len(_proximas)} convocatoria(s) cierran en los próximos 15 días</div>'
            f'{_items}</div>',
            unsafe_allow_html=True
        )


n_conv  = df_c["id"].nunique() if not df_c.empty else 0
n_proy  = df_p["id"].nunique() if not df_p.empty else 0
m_conv  = df_c["Monto"].sum()  if not df_c.empty else 0
v_proy  = df_p["Valor"].sum()  if not df_p.empty else 0
n_ind   = len(df_i)            if not df_i.empty else 0
conv_cp = df_c[df_c["N° proyectos"]>0]["id"].nunique() if not df_c.empty else 0
pct_cp  = round(conv_cp/max(n_conv,1)*100) if n_conv>0 else 0

st.markdown(
    '<div style="background:linear-gradient(135deg,#003d6c 0%,#005931 100%);'
    'border-radius:12px;padding:34px 40px 30px;margin-bottom:24px;'
    'box-shadow:0 6px 15px rgba(0,0,0,.1)">'
    '<div style="font-family:\'DM Serif Display\',serif;font-size:2.2rem;color:#fff;margin:0 0 8px">'
    'Seguimiento de Convocatorias &amp; Proyectos</div>'
    '<div style="color:#a5d6a7;font-size:.9rem;letter-spacing:.02em">'
    'Matriz de seguimiento SDP · Actualización automática cada 5 minutos</div></div>',
    unsafe_allow_html=True)

st.markdown(f"""
<div style="display:flex;gap:14px;margin-bottom:24px;align-items:stretch;flex-wrap:wrap;">
    {kpi("Convocatorias",  n_conv,           "en filtros activos",  style="dark-blue",  flex="1.5")}
    {kpi("Proyectos",      n_proy,           "formulados",          style="dark-green", flex="1.5")}
    {kpi("Con proyectos",  f"{conv_cp}",     f"{pct_cp}% de conv.", border_color="#d88c16", flex="1")}
    {kpi("Monto convoc.",  fmt_money(m_conv),"suma total",          border_color="#cf7000", flex="1")}
    {kpi("Valor proy.",    fmt_money(v_proy),"suma total",          border_color="#47b1d5", flex="1")}
    {kpi("Indicadores MGA",n_ind,            "registros",           border_color="#1754ab", flex="1")}
</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# GENERADOR PDF — Ficha de convocatoria
# ══════════════════════════════════════════════════════════════════════════════
def _gen_pdf_convocatoria(cr, proy_sub, ind_d_local):
    """Genera el PDF de ficha completa de una convocatoria y retorna bytes."""
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
        KeepTogether,
    )
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    W, H = A4
    MARGIN = 2.0 * cm

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title=str(cr.get("Convocatoria","Convocatoria")),
        author="SDP — Secretaría de Planeación",
    )

    # ── Paleta ────────────────────────────────────────────────────────────────
    C_DARK   = colors.HexColor("#003d6c")
    C_GREEN  = colors.HexColor("#005931")
    C_BLUE   = colors.HexColor("#1754ab")
    C_AMBER  = colors.HexColor("#cf7000")
    C_TEAL   = colors.HexColor("#47b1d5")
    C_LIGHT  = colors.HexColor("#f0f6ff")
    C_LGRAY  = colors.HexColor("#f5f5f5")
    C_BORDER = colors.HexColor("#d0dce8")
    C_WHITE  = colors.white
    C_TEXT   = colors.HexColor("#222222")
    C_MUTED  = colors.HexColor("#666666")

    # ── Estilos ───────────────────────────────────────────────────────────────
    def S(name, **kw):
        base = {
            "fontName": "Helvetica", "fontSize": 10,
            "textColor": C_TEXT, "leading": 14,
            "spaceAfter": 0, "spaceBefore": 0,
        }
        base.update(kw)
        return ParagraphStyle(name, **base)

    sTitle   = S("title",   fontName="Helvetica-Bold", fontSize=20, textColor=C_WHITE,   leading=26)
    sSub     = S("sub",     fontName="Helvetica",      fontSize=9,  textColor=colors.HexColor("#a5d6a7"), leading=13)
    sH2      = S("h2",      fontName="Helvetica-Bold", fontSize=13, textColor=C_DARK,    leading=18, spaceBefore=10)
    sH3      = S("h3",      fontName="Helvetica-Bold", fontSize=10, textColor=C_BLUE,    leading=14)
    sLabel   = S("label",   fontName="Helvetica-Bold", fontSize=8,  textColor=C_MUTED,   leading=12)
    sValue   = S("value",   fontName="Helvetica",      fontSize=9,  textColor=C_TEXT,    leading=13)
    sKpiLbl  = S("kpilbl",  fontName="Helvetica-Bold", fontSize=7,  textColor=C_MUTED,   leading=10, alignment=TA_CENTER)
    sKpiVal  = S("kpival",  fontName="Helvetica-Bold", fontSize=16, textColor=C_DARK,    leading=20, alignment=TA_CENTER)
    sSmall   = S("small",   fontName="Helvetica",      fontSize=7,  textColor=C_MUTED,   leading=10, alignment=TA_CENTER)
    sProy    = S("proy",    fontName="Helvetica-Bold", fontSize=9,  textColor=C_BLUE,    leading=13)
    sBody    = S("body",    fontName="Helvetica",      fontSize=9,  textColor=C_TEXT,    leading=14)
    sFooter  = S("footer",  fontName="Helvetica",      fontSize=7,  textColor=C_MUTED,   leading=10, alignment=TA_CENTER)
    sTH      = S("th",      fontName="Helvetica-Bold", fontSize=8,  textColor=C_WHITE,   leading=11, alignment=TA_CENTER)
    sTD      = S("td",      fontName="Helvetica",      fontSize=8,  textColor=C_TEXT,    leading=11)

    story = []
    TW = W - 2 * MARGIN   # usable width

    # ── ENCABEZADO con fondo degradado simulado ───────────────────────────────
    header_data = [[
        Paragraph("SDP", S("sdp", fontName="Helvetica-Bold", fontSize=22, textColor=C_WHITE, leading=26)),
        Paragraph(str(cr.get("Convocatoria", "")), sTitle),
    ]]
    header_tbl = Table(header_data, colWidths=[2.5*cm, TW - 2.5*cm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), C_DARK),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0), (-1,-1), 14),
        ("RIGHTPADDING",(0,0),(-1,-1), 14),
        ("TOPPADDING", (0,0), (-1,-1), 18),
        ("BOTTOMPADDING",(0,0),(-1,-1), 18),
        ("LINEBELOW",  (0,0), (-1,-1), 4, C_GREEN),
    ]))
    story.append(header_tbl)

    # Sub-header: estado + fecha generación
    sub_data = [[
        Paragraph(f"Estado: {cr.get('Estado','—')}  ·  Sector: {cr.get('Sectores','—')}", sSub),
        Paragraph(f"Generado: {date.today().strftime('%d/%m/%Y')}", S("sub_r", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#a5d6a7"), leading=12, alignment=TA_RIGHT)),
    ]]
    sub_tbl = Table(sub_data, colWidths=[TW*0.65, TW*0.35])
    sub_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), C_GREEN),
        ("LEFTPADDING",(0,0), (-1,-1), 14),
        ("RIGHTPADDING",(0,0),(-1,-1), 14),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(sub_tbl)
    story.append(Spacer(1, 14))

    # ── KPIs en fila ──────────────────────────────────────────────────────────
    v_conv  = float(cr.get("Valor proyectos", 0) or 0)
    cob_val = cr.get("Cobertura (%)")
    cob_str = f"{cob_val:.1f}%" if pd.notna(cob_val) else "—"
    n_p_c   = int(cr.get("N° proyectos", 0))

    def kpi_cell(label, value, sub="", bg=C_LIGHT, tc=C_DARK):
        return [
            Paragraph(label, sKpiLbl),
            Paragraph(str(value), S("kv", fontName="Helvetica-Bold", fontSize=15, textColor=tc, leading=19, alignment=TA_CENTER)),
            Paragraph(sub, sSmall),
        ]

    kpi_cols = [
        kpi_cell("MONTO DISPONIBLE",    fmt_money(cr.get("Monto",0)), "convocado"),
        kpi_cell("VALOR FORMULADO",     fmt_money(v_conv),            "en proyectos", tc=C_BLUE),
        kpi_cell("COBERTURA",           cob_str,                      "financiera",   tc=C_AMBER),
        kpi_cell("PROYECTOS",           str(n_p_c),                   "formulados",   tc=C_GREEN),
    ]
    kpi_w = TW / 4
    kpi_tbl = Table(kpi_cols, colWidths=[kpi_w]*4, rowHeights=[12, 22, 10])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), C_LIGHT),
        ("BACKGROUND",    (1,0), (1,-1),  colors.HexColor("#eef4ff")),
        ("BACKGROUND",    (2,0), (2,-1),  colors.HexColor("#fff8ee")),
        ("BACKGROUND",    (3,0), (3,-1),  colors.HexColor("#eefbf3")),
        ("BOX",           (0,0), (0,-1),  0.5, C_BORDER),
        ("BOX",           (1,0), (1,-1),  0.5, C_BORDER),
        ("BOX",           (2,0), (2,-1),  0.5, C_BORDER),
        ("BOX",           (3,0), (3,-1),  0.5, C_BORDER),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 16))

    # ── FECHAS ────────────────────────────────────────────────────────────────
    fecha_data = [[
        Paragraph("FECHA APERTURA", sKpiLbl),
        Paragraph(str(cr.get("Fecha apertura","—")), S("fd", fontName="Helvetica-Bold", fontSize=10, textColor=C_DARK, leading=14, alignment=TA_CENTER)),
        Paragraph("FECHA CIERRE", sKpiLbl),
        Paragraph(str(cr.get("Fecha cierre","—")), S("fc", fontName="Helvetica-Bold", fontSize=10, textColor=C_AMBER, leading=14, alignment=TA_CENTER)),
    ]]
    fecha_tbl = Table(fecha_data, colWidths=[TW*0.15, TW*0.35, TW*0.15, TW*0.35])
    fecha_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), C_LGRAY),
        ("TOPPADDING",    (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("BOX",           (0,0), (-1,-1), 0.5, C_BORDER),
        ("INNERGRID",     (0,0), (-1,-1), 0.3, C_BORDER),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(fecha_tbl)
    story.append(Spacer(1, 18))

    # ── DESCRIPCIÓN ───────────────────────────────────────────────────────────
    def desc_section(title, value):
        if not value or str(value).strip() in ("", "—", "None"): return []
        return [
            Paragraph(title, sH3),
            Spacer(1, 3),
            Paragraph(str(value).strip(), sBody),
            Spacer(1, 10),
        ]

    story += desc_section("Qué ofrece",          cr.get("Qué ofrece",""))
    story += desc_section("Quiénes pueden participar", cr.get("Quiénes participan",""))
    story += desc_section("Público priorizado",   cr.get("Público priorizado",""))

    # ── METADATOS en tabla de 2 columnas ─────────────────────────────────────
    meta_rows = []
    for lbl, val in [
        ("Dependencias",  cr.get("Dependencias","")),
        ("Ubicación",     cr.get("Ubicación","")),
        ("Segmentos",     cr.get("Segmentos","")),
        ("Aliados",       cr.get("Aliados","")),
        ("Contacto",      cr.get("Contacto","")),
    ]:
        v = str(val).strip()
        if v and v not in ("", "—", "None"):
            meta_rows.append([
                Paragraph(lbl, sLabel),
                Paragraph(v, sValue),
            ])

    if meta_rows:
        story.append(Paragraph("Información adicional", sH2))
        story.append(Spacer(1, 6))
        meta_tbl = Table(meta_rows, colWidths=[3*cm, TW - 3*cm])
        meta_tbl.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LINEBELOW",     (0,0), (-1,-1), 0.3, C_BORDER),
            ("BACKGROUND",    (0,0), (0,-1),  C_LGRAY),
            ("LEFTPADDING",   (0,0), (0,-1),  8),
            ("RIGHTPADDING",  (1,0), (1,-1),  4),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 18))

    # ── PROYECTOS ASOCIADOS ───────────────────────────────────────────────────
    if not proy_sub.empty:
        story.append(HRFlowable(width=TW, thickness=1.5, color=C_BLUE, spaceAfter=8))
        story.append(Paragraph(f"Proyectos asociados ({len(proy_sub)})", sH2))
        story.append(Spacer(1, 8))

        # Tabla resumen de proyectos
        th = [Paragraph(h, sTH) for h in ["Proyecto","Dependencia","Valor","Contrapartida","Beneficiarios","BPIN"]]
        rows = [th]
        for _, pr in proy_sub.iterrows():
            rows.append([
                Paragraph(str(pr.get("Proyecto",""))[:70], sTD),
                Paragraph(str(pr.get("Dependencia","—")), sTD),
                Paragraph(fmt_money(pr.get("Valor",0)), sTD),
                Paragraph(fmt_money(pr.get("Contrapartida",0)), sTD),
                Paragraph(str(int(pr.get("Total beneficiarios",0))), sTD),
                Paragraph(str(pr.get("BPIN","—")), sTD),
            ])
        col_w = [TW*0.30, TW*0.22, TW*0.12, TW*0.12, TW*0.12, TW*0.12]
        p_tbl = Table(rows, colWidths=col_w, repeatRows=1)
        p_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  C_DARK),
            ("BACKGROUND",    (0,1), (-1,-1), C_WHITE),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_LGRAY]),
            ("GRID",          (0,0), (-1,-1), 0.3, C_BORDER),
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ]))
        story.append(KeepTogether([p_tbl]))
        story.append(Spacer(1, 14))

        # Detalle de indicadores por proyecto
        for _, pr in proy_sub.iterrows():
            pid_pdf = int(pr["id"])
            inds = ind_d_local.get(pid_pdf, [])
            if not inds: continue
            proy_block = [
                Spacer(1, 6),
                Paragraph(str(pr.get("Proyecto",""))[:90], sProy),
                Spacer(1, 4),
            ]
            ind_th = [Paragraph(h, sTH) for h in ["Codigo","Indicador","Vigencia","Meta Proyecto"]]
            ind_rows = [ind_th]
            for ind in inds:
                ind_rows.append([
                    Paragraph(str(ind.get("codigo","")), sTD),
                    Paragraph(str(ind.get("nombre",""))[:80], sTD),
                    Paragraph(str(ind.get("vigencia","")), sTD),
                    Paragraph(str(ind.get("meta_proyecto","")), sTD),
                ])
            ind_tbl = Table(ind_rows, colWidths=[TW*0.15, TW*0.50, TW*0.17, TW*0.18], repeatRows=1)
            ind_tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,0),  C_BLUE),
                ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_LGRAY]),
                ("GRID",          (0,0), (-1,-1), 0.3, C_BORDER),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING",   (0,0), (-1,-1), 5),
                ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ]))
            proy_block.append(ind_tbl)
            story.append(KeepTogether(proy_block))

    # ── FOOTER ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width=TW, thickness=0.5, color=C_BORDER))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Secretaría de Planeación (SDP) · Seguimiento de Convocatorias y Proyectos · "
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        sFooter
    ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# TABS — 4 pestañas, sin Trazabilidad
# ══════════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Convocatorias", "Proyectos", "Mapa", "Asistente IA", "Exportar",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 · CONVOCATORIAS
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    st.markdown(sec_title("Convocatorias",
        "Selecciona una convocatoria para ver su ficha completa, o explora el resumen general"),
        unsafe_allow_html=True)

    # ── Búsqueda por palabra clave ──────────────────────────────────────────
    _sc1, _sc2 = st.columns([3, 2])
    with _sc1:
        busq_c = st.text_input(
            "Buscar convocatoria",
            placeholder="Buscar por nombre, sector, estado, dependencia...",
            key="busq_c", label_visibility="collapsed"
        )
    with _sc2:
        if busq_c.strip():
            _term_c = busq_c.strip().lower()
            _cols_c = ["Convocatoria","Estado","Sectores","Segmentos",
                       "Ubicación","Dependencias","Aliados","Contacto",
                       "Qué ofrece","Quiénes participan","Público priorizado"]
            _cols_c = [c for c in _cols_c if c in df_c.columns]
            _mask_c = df_c[_cols_c].apply(
                lambda col: col.astype(str).str.lower().str.contains(_term_c, na=False)
            ).any(axis=1)
            df_c_busq = df_c[_mask_c]
            n_res_c   = len(df_c_busq)
            st.markdown(
                f'<div style="padding:10px 0;font-size:.82rem;color:#1754ab;font-weight:600">'
                f'{n_res_c} resultado(s) para "{busq_c}"</div>',
                unsafe_allow_html=True)
        else:
            df_c_busq = df_c
            st.markdown(
                f'<div style="padding:10px 0;font-size:.82rem;color:#888">'
                f'{len(df_c)} convocatoria(s) · usa el buscador para filtrar</div>',
                unsafe_allow_html=True)

    opciones_c = ["— Ver resumen general —"] + (
        sorted(df_c_busq["Convocatoria"].tolist()) if not df_c_busq.empty else [])
    sel_c = st.selectbox("Convocatoria", opciones_c, key="sel_c",
                          label_visibility="collapsed")

    # ══ MODO A · RESUMEN GENERAL ═════════════════════════════════════════════
    if sel_c == "— Ver resumen general —":
        if not df_c.empty:
            # Gráficas fila 1
            ca, cb = st.columns([3, 2])
            with ca:
                st.markdown(bar_chart(df_c["Estado"].value_counts(), "Por estado"),
                            unsafe_allow_html=True)
                sec_exp = df_c["Sectores"].str.split(" · ").explode().str.strip().value_counts()
                sec_exp = sec_exp[sec_exp.index.str.len()>0]
                if not sec_exp.empty:
                    st.markdown(bar_chart(sec_exp, "Por sector", max_bars=15),
                                unsafe_allow_html=True)
                dep_exp = df_c["Dependencias"].str.split(" · ").explode().str.strip().value_counts()
                dep_exp = dep_exp[dep_exp.index.str.len()>0]
                if not dep_exp.empty:
                    st.markdown(bar_chart(dep_exp, "Por dependencia", max_bars=12),
                                unsafe_allow_html=True)
            with cb:
                seg_exp = df_c["Segmentos"].str.split(" · ").explode().str.strip().value_counts()
                seg_exp = seg_exp[seg_exp.index.str.len()>0]
                if not seg_exp.empty:
                    st.markdown(donut_chart(seg_exp, "Por segmento"), unsafe_allow_html=True)
                ubi_exp = df_c["Ubicación"].str.split(" · ").explode().str.strip().value_counts()
                ubi_exp = ubi_exp[ubi_exp.index.str.len()>0]
                if not ubi_exp.empty:
                    st.markdown(donut_chart(ubi_exp, "Por ubicación"), unsafe_allow_html=True)
                st.markdown(bar_chart(
                    df_c.groupby("Estado")["Monto"].sum().sort_values(ascending=False),
                    "Monto total por estado", fmt_val=fmt_money), unsafe_allow_html=True)

            # Gráficas fila 2 — análisis financiero (venía de Trazabilidad)
            st.markdown(sec_title("Análisis financiero",
                "Relación entre montos de convocatorias y valor de proyectos formulados"),
                unsafe_allow_html=True)
            fa, fb = st.columns([3, 2])
            with fa:
                if "N° proyectos" in df_c.columns:
                    st.markdown(bar_chart(
                        df_c.set_index("Convocatoria")["N° proyectos"].sort_values(ascending=False),
                        "Proyectos formulados por convocatoria", max_bars=20),
                        unsafe_allow_html=True)
                if "Valor proyectos" in df_c.columns:
                    st.markdown(bar_chart(
                        df_c.set_index("Convocatoria")["Valor proyectos"].sort_values(ascending=False).head(12),
                        "Valor total formulado por convocatoria (top 12)", fmt_val=fmt_money),
                        unsafe_allow_html=True)
            with fb:
                if "Cobertura (%)" in df_c.columns:
                    cob_vals = df_c["Cobertura (%)"].dropna()
                    if not cob_vals.empty:
                        cob_r = pd.cut(cob_vals, bins=[0,25,50,75,100,float("inf")],
                                       labels=["0–25%","25–50%","50–75%","75–100%",">100%"])
                        st.markdown(donut_chart(cob_r.value_counts(),
                            "Distribución cobertura financiera"), unsafe_allow_html=True)
                        # Top convocatorias por cobertura
                        top_cob = df_c[df_c["Cobertura (%)"].notna()].nlargest(8,"Cobertura (%)")
                        if not top_cob.empty:
                            st.markdown(bar_chart(
                                top_cob.set_index("Convocatoria")["Cobertura (%)"].sort_values(ascending=False),
                                "Top cobertura financiera (%)"),
                                unsafe_allow_html=True)

            # Alerta: convocatorias sin proyectos
            sin = df_c[df_c["N° proyectos"]==0] if not df_c.empty else pd.DataFrame()
            if not sin.empty:
                with st.expander(f"{len(sin)} convocatoria(s) sin proyectos asociados"):
                    sc_cols = ["Convocatoria","Estado","Monto","Sectores","Fecha cierre"]
                    sc_cols = [c for c in sc_cols if c in sin.columns]
                    st.dataframe(sin[sc_cols].reset_index(drop=True),
                        use_container_width=True, hide_index=True,
                        column_config={"Monto": st.column_config.NumberColumn("Monto $", format="$ %d")})

        st.markdown(sec_title("Directorio de convocatorias"), unsafe_allow_html=True)
        lc = ["Convocatoria","Estado","Fecha apertura","Fecha cierre","Monto",
              "Sectores","Segmentos","Ubicación","N° proyectos","Cobertura (%)","Contacto"]
        lc = [c for c in lc if c in df_c_busq.columns]
        st.dataframe(df_c_busq[lc].reset_index(drop=True), use_container_width=True, height=420,
            hide_index=True,
            column_config={
                "Convocatoria":  st.column_config.TextColumn(width=280),
                "Monto":         st.column_config.NumberColumn("Monto $",    format="$ %d"),
                "N° proyectos":  st.column_config.NumberColumn("Proyectos",  width=90),
                "Cobertura (%)": st.column_config.NumberColumn("Cob. %",     format="%.1f%%"),
            })

    # ══ MODO B · FICHA DE CONVOCATORIA ═══════════════════════════════════════
    else:
        row_c = df_c[df_c["Convocatoria"]==sel_c]
        if row_c.empty:
            st.markdown(empty_state("Convocatoria no encontrada en los filtros activos."),
                        unsafe_allow_html=True)
        else:
            cr = row_c.iloc[0]

            # Badges
            estado_color = "#17743d" if "vigente" in str(cr["Estado"]).lower() else "#cf7000"
            bdgs = badge(cr["Estado"], estado_color)
            for s in str(cr.get("Sectores","")).split(" · "):
                if s.strip(): bdgs += badge(s.strip(), "#1754ab")
            for s in str(cr.get("Segmentos","")).split(" · "):
                if s.strip(): bdgs += badge(s.strip(), "#47b1d5")

            # KPIs de cobertura
            v_conv   = float(cr.get("Valor proyectos", 0) or 0)
            cob_conv = cr.get("Cobertura (%)")
            cob_str  = f"{cob_conv:.1f}%" if pd.notna(cob_conv) else "—"
            n_proy_c = int(cr["N° proyectos"])

            st.markdown(f"""
<div style="background:#f8fbff;border:1px solid #cce0f5;border-left:5px solid #1754ab;
border-radius:10px;padding:26px 30px;margin:14px 0 20px">
  <div style="font-family:'DM Serif Display',serif;font-size:1.55rem;color:#003d6c;margin-bottom:10px">
    {cr['Convocatoria']}</div>
  <div style="margin-bottom:18px">{bdgs}</div>
  {stat_grid(
      ("Monto disponible",      fmt_money(cr['Monto']),  "#005931"),
      ("Valor formulado",       fmt_money(v_conv),        "#1754ab"),
      ("Cobertura financiera",  cob_str,                  "#cf7000"),
      ("Proyectos formulados",  str(n_proy_c),            "#003d6c"),
      ("Fecha apertura",        cr['Fecha apertura'],     "#444"),
      ("Fecha cierre",          cr['Fecha cierre'],       "#444"),
  )}
  {field_row("Qué ofrece",         cr.get("Qué ofrece",""))}
  {field_row("Quiénes participan", cr.get("Quiénes participan",""))}
  {field_row("Público priorizado", cr.get("Público priorizado",""))}
  {field_row("Ubicación",          cr.get("Ubicación",""))}
  {field_row("Dependencias",       cr.get("Dependencias",""))}
  {field_row("Aliados",            cr.get("Aliados",""))}
  {field_row("Contacto",           cr.get("Contacto",""))}
</div>""", unsafe_allow_html=True)

            # Botón descarga PDF
            _pdf_bytes = _gen_pdf_convocatoria(cr, df_p[df_p["convocatoria_id"]==int(cr["id"])], _ind_d)
            _pdf_name  = re.sub(r"[^a-zA-Z0-9_]","_", str(cr["Convocatoria"]))[:60] + ".pdf"
            st.download_button(
                label="Descargar ficha en PDF",
                data=_pdf_bytes,
                file_name=_pdf_name,
                mime="application/pdf",
                use_container_width=False,
            )

            # Proyectos asociados como expanders
            proy_sub = df_p[df_p["convocatoria_id"]==int(cr["id"])]
            if proy_sub.empty:
                st.markdown(empty_state("Esta convocatoria no tiene proyectos formulados aún."),
                            unsafe_allow_html=True)
            else:
                st.markdown(sec_title("Proyectos asociados",
                    f"{len(proy_sub)} proyecto(s) · expande cada tarjeta para ver ficha completa"),
                    unsafe_allow_html=True)

                # Mini-resumen de proyectos: barras de valor
                if len(proy_sub) > 1:
                    st.markdown(bar_chart(
                        proy_sub.set_index("Proyecto")["Valor"].sort_values(ascending=False),
                        "Valor por proyecto", fmt_val=fmt_money, max_bars=15),
                        unsafe_allow_html=True)

                for _, pr in proy_sub.iterrows():
                    pid_p = int(pr["id"])
                    cob_p = df_r.loc[df_r["id"]==pid_p,"Cobertura (%)"].values
                    cob_p_str = f"{cob_p[0]:.1f}%" if len(cob_p)>0 and pd.notna(cob_p[0]) else "—"

                    with st.expander(
                        f"**{pr['Proyecto']}** · {pr['Dependencia']} · "
                        f"{fmt_money(pr['Valor'])} · Cob. {cob_p_str}"
                    ):
                        e1,e2,e3,e4 = st.columns(4)
                        e1.metric("Valor",         fmt_money(pr["Valor"]))
                        e2.metric("Contrapartida", fmt_money(pr.get("Contrapartida",0)))
                        e3.metric("Beneficiarios", int(pr.get("Total beneficiarios",0)))
                        e4.metric("BPIN",          pr.get("BPIN","—"))

                        fh = ""
                        for f,lbl in [("Responsable","Responsable"),
                                      ("Municipios","Municipios"),
                                      ("Tipos beneficiarios","Tipos de beneficiarios")]:
                            v = pr.get(f,"")
                            if v and str(v).strip() not in ("","—","0","None"):
                                fh += field_row(lbl, v)
                        if fh:
                            st.markdown(
                                f'<div style="background:#fafafa;border:1px solid #ececec;'
                                f'border-radius:8px;padding:14px 18px;margin:10px 0">'
                                f'{fh}</div>', unsafe_allow_html=True)

                        inds_p = _ind_d.get(pid_p, [])
                        if inds_p:
                            st.markdown(
                                f'<div style="font-size:.78rem;font-weight:600;color:#1754ab;'
                                f'margin:14px 0 6px">{len(inds_p)} indicador(es) MGA</div>',
                                unsafe_allow_html=True)
                            _ind_table(inds_p)
                        else:
                            st.caption("Sin indicadores MGA registrados.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 · PROYECTOS
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    st.markdown(sec_title("Proyectos",
        "Selecciona un proyecto para ver su ficha completa, o explora el resumen general"),
        unsafe_allow_html=True)

    # ── Búsqueda por palabra clave ──────────────────────────────────────────
    _sp1, _sp2 = st.columns([3, 2])
    with _sp1:
        busq_p = st.text_input(
            "Buscar proyecto",
            placeholder="Buscar por nombre, BPIN, dependencia, municipio, responsable...",
            key="busq_p", label_visibility="collapsed"
        )
    with _sp2:
        if busq_p.strip():
            _term_p = busq_p.strip().lower()
            _cols_p = ["Proyecto","BPIN","Dependencia","Responsable",
                       "Municipios","Tipos beneficiarios","Indicadores MGA"]
            _cols_p = [c for c in _cols_p if c in df_p.columns]
            _mask_p = df_p[_cols_p].apply(
                lambda col: col.astype(str).str.lower().str.contains(_term_p, na=False)
            ).any(axis=1)
            df_p_busq = df_p[_mask_p]
            n_res_p   = len(df_p_busq)
            st.markdown(
                f'<div style="padding:10px 0;font-size:.82rem;color:#17743d;font-weight:600">'
                f'{n_res_p} resultado(s) para "{busq_p}"</div>',
                unsafe_allow_html=True)
        else:
            df_p_busq = df_p
            st.markdown(
                f'<div style="padding:10px 0;font-size:.82rem;color:#888">'
                f'{len(df_p)} proyecto(s) · usa el buscador para filtrar</div>',
                unsafe_allow_html=True)

    opciones_p = ["— Ver resumen general —"] + (
        sorted(df_p_busq["Proyecto"].tolist()) if not df_p_busq.empty else [])
    sel_p = st.selectbox("Proyecto", opciones_p, key="sel_p",
                          label_visibility="collapsed")

    # ══ MODO A · RESUMEN GENERAL ═════════════════════════════════════════════
    if sel_p == "— Ver resumen general —":
        if not df_p.empty:
            pa, pb = st.columns([3, 2])
            with pa:
                st.markdown(bar_chart(df_p["Dependencia"].value_counts(),
                    "Por dependencia", max_bars=15), unsafe_allow_html=True)
                st.markdown(bar_chart(
                    df_p.nlargest(15,"Valor").set_index("Proyecto")["Valor"],
                    "Top 15 por valor", fmt_val=fmt_money), unsafe_allow_html=True)
                if not df_r.empty:
                    vxd = df_r.groupby("Dependencia")["Valor"].sum().sort_values(ascending=False)
                    st.markdown(bar_chart(vxd, "Valor total formulado por dependencia",
                        fmt_val=fmt_money, max_bars=15), unsafe_allow_html=True)
            with pb:
                st.markdown(donut_chart(df_p["Responsable"].value_counts(),
                    "Por responsable"), unsafe_allow_html=True)
                mun_exp = df_p["Municipios"].str.split(" · ").explode().str.strip().value_counts()
                mun_exp = mun_exp[mun_exp.index.str.len()>0]
                if not mun_exp.empty:
                    st.markdown(donut_chart(mun_exp, "Cobertura municipal"),
                                unsafe_allow_html=True)
                ben_dep = df_p.groupby("Dependencia")["Total beneficiarios"].sum()
                ben_dep = ben_dep[ben_dep>0]
                if not ben_dep.empty:
                    st.markdown(bar_chart(ben_dep.sort_values(ascending=False),
                        "Beneficiarios por dependencia"), unsafe_allow_html=True)

            # Fila 2 — indicadores MGA
            if not df_i.empty:
                st.markdown(sec_title("Indicadores MGA",
                    "Análisis de metas físicas e indicadores estándar"),
                    unsafe_allow_html=True)
                ma, mb = st.columns([3, 2])
                with ma:
                    st.markdown(bar_chart(df_i["nombre"].value_counts().head(15),
                        "Indicadores más usados"), unsafe_allow_html=True)
                    meta_sum = df_i.groupby("nombre")["meta_proyecto"].sum() \
                                   .sort_values(ascending=False).head(12)
                    meta_sum = meta_sum[meta_sum>0]
                    if not meta_sum.empty:
                        st.markdown(bar_chart(meta_sum, "Meta total por indicador"),
                                    unsafe_allow_html=True)
                with mb:
                    st.markdown(donut_chart(
                        df_i["vigencia"].astype(str).value_counts(),
                        "Distribución por vigencia"), unsafe_allow_html=True)
                    ixp = df_i.groupby("Proyecto")["codigo"].count() \
                               .sort_values(ascending=False).head(10)
                    st.markdown(bar_chart(ixp, "Top 10 proyectos con más indicadores"),
                                unsafe_allow_html=True)

        st.markdown(sec_title("Directorio de proyectos"), unsafe_allow_html=True)
        ps = ["Proyecto","BPIN","Valor","Contrapartida","Dependencia","Responsable",
              "Municipios","Total beneficiarios","N° indicadores MGA"]
        ps = [c for c in ps if c in df_p_busq.columns]
        st.dataframe(df_p_busq[ps].reset_index(drop=True), use_container_width=True, height=420,
            hide_index=True,
            column_config={
                "Proyecto":            st.column_config.TextColumn(width=280),
                "Valor":               st.column_config.NumberColumn("Valor $",       format="$ %d"),
                "Contrapartida":       st.column_config.NumberColumn("Contrapartida", format="$ %d"),
                "Total beneficiarios": st.column_config.NumberColumn("Beneficiarios", width=110),
                "N° indicadores MGA":  st.column_config.NumberColumn("Indicadores",   width=100),
            })

    # ══ MODO B · FICHA DE PROYECTO ════════════════════════════════════════════
    else:
        rp_match = df_p[df_p["Proyecto"]==sel_p]
        if rp_match.empty:
            st.markdown(empty_state("Proyecto no encontrado en los filtros activos."),
                        unsafe_allow_html=True)
        else:
            rp     = rp_match.iloc[0]
            pid_sel= int(rp["id"])

            # Convocatoria vinculada
            conv_id   = rp.get("convocatoria_id")
            conv_link = df_conv[df_conv["id"]==pd.to_numeric(conv_id, errors="coerce")] if conv_id else pd.DataFrame()
            conv_row  = conv_link.iloc[0] if not conv_link.empty else None

            # Cobertura de este proyecto
            cob_proy = df_r.loc[df_r["id"]==pid_sel,"Cobertura (%)"].values
            cob_proy_str = f"{cob_proy[0]:.1f}%" if len(cob_proy)>0 and pd.notna(cob_proy[0]) else "—"

            # Badges
            bdgs_p = badge(rp["Dependencia"], "#1754ab")
            if conv_row is not None:
                est_c = "#17743d" if "vigente" in str(conv_row["Estado"]).lower() else "#cf7000"
                bdgs_p += badge(conv_row["Estado"], est_c)
            for m in str(rp.get("Municipios","")).split(" · "):
                if m.strip(): bdgs_p += badge(m.strip(), "#47b1d5")

            st.markdown(f"""
<div style="background:#f8fff9;border:1px solid #b8dfc4;border-left:5px solid #17743d;
border-radius:10px;padding:26px 30px;margin:14px 0 20px">
  <div style="font-family:'DM Serif Display',serif;font-size:1.55rem;color:#003d6c;margin-bottom:10px">
    {rp['Proyecto']}</div>
  <div style="margin-bottom:18px">{bdgs_p}</div>
  {stat_grid(
      ("Valor del proyecto",   fmt_money(rp['Valor']),                  "#005931"),
      ("Contrapartida",        fmt_money(rp.get('Contrapartida',0)),    "#444"),
      ("Cobertura convoc.",    cob_proy_str,                            "#cf7000"),
      ("Total beneficiarios",  str(int(rp.get('Total beneficiarios',0))), "#1754ab"),
      ("BPIN",                 rp.get('BPIN','—'),                      "#444"),
      ("Responsable",          rp.get('Responsable','—'),               "#444"),
  )}
  {field_row("Dependencia",         rp.get("Dependencia",""))}
  {field_row("Municipios",          rp.get("Municipios",""))}
  {field_row("Tipos beneficiarios", rp.get("Tipos beneficiarios",""))}
  {field_row("Indicadores MGA",     rp.get("Indicadores MGA",""))}
</div>""", unsafe_allow_html=True)

            # Card de convocatoria vinculada
            if conv_row is not None:
                mc_v   = float(conv_row.get("Monto",0) or 0)
                est_c  = "#17743d" if "vigente" in str(conv_row["Estado"]).lower() else "#cf7000"
                st.markdown(f"""
<div style="background:#f8fbff;border:1px solid #cce0f5;border-left:5px solid #1754ab;
border-radius:10px;padding:20px 26px;margin-bottom:20px">
  <div style="font-size:.68rem;color:#888;text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px">
    Convocatoria vinculada</div>
  <div style="font-family:'DM Serif Display',serif;font-size:1.2rem;color:#003d6c;margin-bottom:12px">
    {conv_row['Convocatoria']}</div>
  {stat_grid(
      ("Estado",         conv_row['Estado'],                     est_c),
      ("Monto total",    fmt_money(mc_v),                        "#005931"),
      ("Fecha apertura", conv_row.get('Fecha apertura','—'),     "#444"),
      ("Fecha cierre",   conv_row.get('Fecha cierre','—'),       "#444"),
  )}
  {field_row("Qué ofrece",         conv_row.get("Qué ofrece",""))}
  {field_row("Quiénes participan", conv_row.get("Quiénes participan",""))}
  {field_row("Público priorizado", conv_row.get("Público priorizado",""))}
  {field_row("Sectores",           conv_row.get("Sectores",""))}
  {field_row("Segmentos",          conv_row.get("Segmentos",""))}
  {field_row("Ubicación",          conv_row.get("Ubicación",""))}
  {field_row("Aliados",            conv_row.get("Aliados",""))}
  {field_row("Contacto",           conv_row.get("Contacto",""))}
</div>""", unsafe_allow_html=True)

            # Indicadores MGA del proyecto
            inds_sel = _ind_d.get(pid_sel, [])
            if inds_sel:
                st.markdown(sec_title("Indicadores MGA",
                    f"{len(inds_sel)} indicador(es) asociados a este proyecto"),
                    unsafe_allow_html=True)
                _ind_table(inds_sel)
            else:
                st.markdown(empty_state("Sin indicadores MGA registrados para este proyecto."),
                            unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 · MAPA — Cobertura municipal Sucre
# ─────────────────────────────────────────────────────────────────────────────
with tab3:
    st.markdown(sec_title("Mapa de cobertura municipal",
        "Proyectos y convocatorias por municipio · Departamento de Sucre"),
        unsafe_allow_html=True)

    # Coordenadas centroide de cada municipio de Sucre (WGS84)
    SUCRE_COORDS = {
        "Sincelejo":          (9.3047,  -75.3978),
        "Buenavista":         (9.3167,  -75.0000),
        "Caimito":            (8.7833,  -75.0833),
        "Colosó":             (9.4833,  -75.3500),
        "Corozal":            (9.3167,  -75.2833),
        "Coveñas":            (9.4000,  -75.6833),
        "Chalán":             (9.5167,  -75.2500),
        "El Roble":           (9.2167,  -75.1833),
        "Galeras":            (9.1167,  -74.8833),
        "Guaranda":           (8.4667,  -74.5333),
        "La Unión":           (8.8667,  -75.2833),
        "Los Palmitos":       (9.3667,  -75.1667),
        "Majagual":           (8.5333,  -74.6333),
        "Morroa":             (9.3333,  -75.3000),
        "Ovejas":             (9.5167,  -75.2167),
        "Palmito":            (9.3833,  -75.5333),
        "Sampués":            (9.1833,  -75.3833),
        "San Benito Abad":    (8.9333,  -75.0167),
        "San Juan de Betulia":(9.2833,  -75.1333),
        "San Marcos":         (8.6667,  -75.1333),
        "San Onofre":         (9.7333,  -75.5167),
        "San Pedro":          (9.3500,  -75.2500),
        "Sincé":              (9.2333,  -75.1500),
        "Sucre":              (8.8167,  -74.7167),
        "Tolú":               (9.5167,  -75.5833),
        "Toluviejo":          (9.4667,  -75.4333),
    }

    # Alias de normalización (por si en BD viene con variante ortográfica)
    ALIAS = {
        "toluuviejo": "Toluviejo", "tolú viejo": "Toluviejo", "tolu viejo": "Toluviejo",
        "san juan betulia": "San Juan de Betulia",
        "santiago de tolú": "Tolú", "santiago de tolu": "Tolú",
        "tolu": "Tolú",
    }

    def _norm_mun(name):
        n = name.strip()
        low = n.lower()
        if low in ALIAS: return ALIAS[low]
        for k in SUCRE_COORDS:
            if k.lower() == low: return k
        return n

    if not _FOLIUM_OK:
        st.warning("Para ver el mapa instala: `pip install folium streamlit-folium`")
    elif df_p.empty and df_c.empty:
        st.markdown(empty_state("Sin datos para mostrar en el mapa."), unsafe_allow_html=True)
    else:
        # ── Construir tabla municipio → proyectos / convocatorias ────────────
        mun_data: dict = {}   # {municipio: {proyectos:[], convocatorias:set()}}

        # Proyectos por municipio
        for _, row in df_p.iterrows():
            muns = [_norm_mun(m) for m in str(row.get("Municipios","")).split(" · ") if m.strip()]
            for m in muns:
                if m not in mun_data:
                    mun_data[m] = {"proyectos": [], "convocatorias": set(), "valor": 0, "beneficiarios": 0}
                mun_data[m]["proyectos"].append(row["Proyecto"])
                mun_data[m]["valor"]        += float(row.get("Valor",0) or 0)
                mun_data[m]["beneficiarios"]+= int(row.get("Total beneficiarios",0) or 0)
                cid = row.get("convocatoria_id")
                if cid:
                    conv_match = df_conv[df_conv["id"]==cid]
                    if not conv_match.empty:
                        mun_data[m]["convocatorias"].add(conv_match.iloc[0]["Convocatoria"])

        # ── Controles fila 1 ─────────────────────────────────────────────────
        mc1, mc2, mc3 = st.columns([2, 2, 3])
        with mc1:
            mapa_capas = st.multiselect(
                "Mostrar en el mapa",
                ["Proyectos", "Convocatorias"],
                default=["Proyectos", "Convocatorias"],
                key="mapa_capas"
            )
            mapa_zoom = st.slider("Zoom inicial", 8, 12, 9, key="mapa_zoom")

        with mc2:
            # Filtro por dependencia
            deps_en_mapa = sorted(df_p["Dependencia"].dropna().unique().tolist()) if not df_p.empty else []
            mapa_dep = st.multiselect(
                "Filtrar por dependencia",
                deps_en_mapa,
                placeholder="Todas las dependencias",
                key="mapa_dep"
            )
            # Búsqueda de municipio
            mapa_busq = st.text_input(
                "Buscar municipio",
                placeholder="Ej: Sincelejo, Corozal...",
                key="mapa_busq"
            )

        with mc3:
            # KPIs rápidos del mapa
            n_mun_cub = sum(1 for m in mun_data if m in SUCRE_COORDS)
            tot_val_map = sum(d["valor"] for d in mun_data.values())
            tot_ben_map = sum(d["beneficiarios"] for d in mun_data.values())
            st.markdown(f"""
<div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:4px">
    {kpi("Municipios cubiertos", f"{n_mun_cub}/26", "del departamento", border_color="#17743d", flex="1")}
    {kpi("Valor total",  fmt_money(tot_val_map), "en municipios", border_color="#1754ab", flex="1")}
    {kpi("Beneficiarios", f"{tot_ben_map:,}", "en municipios", border_color="#cf7000", flex="1")}
</div>""", unsafe_allow_html=True)

        # Aplicar filtro de dependencia a mun_data
        if mapa_dep:
            df_p_mapa = df_p[df_p["Dependencia"].isin(mapa_dep)]
            mun_data_filtrado: dict = {}
            for _, row in df_p_mapa.iterrows():
                muns = [_norm_mun(mn) for mn in str(row.get("Municipios","")).split(" · ") if mn.strip()]
                for mn in muns:
                    if mn not in mun_data_filtrado:
                        mun_data_filtrado[mn] = {"proyectos":[], "convocatorias":set(), "valor":0, "beneficiarios":0}
                    mun_data_filtrado[mn]["proyectos"].append(row["Proyecto"])
                    mun_data_filtrado[mn]["valor"]        += float(row.get("Valor",0) or 0)
                    mun_data_filtrado[mn]["beneficiarios"]+= int(row.get("Total beneficiarios",0) or 0)
                    cid2 = row.get("convocatoria_id")
                    if cid2:
                        cm2 = df_conv[df_conv["id"]==cid2]
                        if not cm2.empty:
                            mun_data_filtrado[mn]["convocatorias"].add(cm2.iloc[0]["Convocatoria"])
            mun_data_render = mun_data_filtrado
        else:
            mun_data_render = mun_data

        # Municipio buscado (para centrar/resaltar)
        mun_busq_norm = _norm_mun(mapa_busq.strip()) if mapa_busq.strip() else None
        mun_busq_found = mun_busq_norm in SUCRE_COORDS if mun_busq_norm else False

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Construir mapa Folium ─────────────────────────────────────────────
        # Centro: municipio buscado si existe, sino centro de Sucre
        _map_center = list(SUCRE_COORDS[mun_busq_norm]) if mun_busq_found else [9.05, -75.15]
        _map_zoom   = 11 if mun_busq_found else mapa_zoom

        m = folium.Map(
            location=_map_center,
            zoom_start=_map_zoom,
            tiles="CartoDB positron",
            control_scale=True,
        )

        # Dibujar todos los municipios como círculos de fondo (gris si sin datos)
        for mun, (lat, lng) in SUCRE_COORDS.items():
            tiene_datos = mun in mun_data_render
            es_buscado  = (mun == mun_busq_norm)
            folium.CircleMarker(
                location=[lat, lng],
                radius=10 if es_buscado else 8,
                color="#e68878" if es_buscado else ("#cccccc" if not tiene_datos else "#003d6c"),
                fill=True,
                fill_color="#e68878" if es_buscado else ("#eeeeee" if not tiene_datos else "#e8f0fe"),
                fill_opacity=0.7 if es_buscado else 0.5,
                weight=3 if es_buscado else 1,
                tooltip=mun if not tiene_datos else None,
            ).add_to(m)

        # ── Ruta entre municipios cubiertos ──────────────────────────────────
        muns_con_proy = [mn for mn in mun_data_render if mn in SUCRE_COORDS]
        if len(muns_con_proy) >= 2:
            # Ordenar de norte a sur (por latitud desc) para trazar ruta coherente
            muns_ordenados = sorted(
                muns_con_proy,
                key=lambda mn: -SUCRE_COORDS[mn][0]
            )
            ruta_coords = [list(SUCRE_COORDS[mn]) for mn in muns_ordenados]
            folium.PolyLine(
                locations=ruta_coords,
                color="#47b1d5",
                weight=2,
                opacity=0.5,
                dash_array="6 4",
                tooltip="Ruta entre municipios cubiertos",
            ).add_to(m)

        # Marcadores con datos
        for mun, data in mun_data_render.items():
            if mun not in SUCRE_COORDS:
                continue
            lat, lng = SUCRE_COORDS[mun]
            n_p   = len(data["proyectos"])
            n_c   = len(data["convocatorias"])
            valor = fmt_money(data["valor"])
            bens  = f'{data["beneficiarios"]:,}'

            # Color según número de proyectos
            if n_p == 0:   color = "#47b1d5"
            elif n_p <= 2: color = "#17743d"
            elif n_p <= 5: color = "#1754ab"
            else:          color = "#003d6c"

            # Radio proporcional al valor
            radius = 10 + min(n_p * 4, 30)

            # Popup HTML rico
            proy_list_html = "".join(
                f'<li style="font-size:.78rem;color:#333;margin-bottom:3px">{p}</li>'
                for p in data["proyectos"][:8]
            )
            if len(data["proyectos"]) > 8:
                proy_list_html += f'<li style="color:#888;font-size:.75rem">...y {len(data["proyectos"])-8} más</li>'

            conv_list_html = "".join(
                f'<li style="font-size:.78rem;color:#1754ab;margin-bottom:3px">{c}</li>'
                for c in list(data["convocatorias"])[:5]
            )

            show_proy = "Proyectos" in mapa_capas
            show_conv = "Convocatorias" in mapa_capas

            popup_html = f"""
<div style="font-family:Arial,sans-serif;min-width:240px;max-width:320px">
  <div style="background:#003d6c;color:#fff;padding:10px 14px;border-radius:6px 6px 0 0;
              font-weight:700;font-size:.95rem">{mun}</div>
  <div style="padding:12px 14px;border:1px solid #e0e0e0;border-top:none;border-radius:0 0 6px 6px">
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:10px">
      <div style="background:#f0f8ff;border-radius:6px;padding:6px 10px;text-align:center">
        <div style="font-size:1.3rem;font-weight:700;color:#003d6c">{n_p}</div>
        <div style="font-size:.68rem;color:#666">Proyectos</div>
      </div>
      <div style="background:#f0fff4;border-radius:6px;padding:6px 10px;text-align:center">
        <div style="font-size:1.3rem;font-weight:700;color:#17743d">{n_c}</div>
        <div style="font-size:.68rem;color:#666">Convocatorias</div>
      </div>
    </div>
    <div style="font-size:.75rem;color:#555;margin-bottom:4px">
      <b>Valor formulado:</b> {valor}
    </div>
    <div style="font-size:.75rem;color:#555;margin-bottom:10px">
      <b>Beneficiarios:</b> {bens}
    </div>
    {'<div style="font-size:.78rem;font-weight:700;color:#003d6c;margin-bottom:4px"> Proyectos</div><ul style="margin:0;padding-left:16px">' + proy_list_html + '</ul>' if show_proy and n_p > 0 else ""}
    {'<div style="font-size:.78rem;font-weight:700;color:#1754ab;margin:8px 0 4px">Convocatorias</div><ul style="margin:0;padding-left:16px">' + conv_list_html + '</ul>' if show_conv and n_c > 0 else ""}
  </div>
</div>"""

            folium.CircleMarker(
                location=[lat, lng],
                radius=radius,
                color=color,
                fill=True,
                fill_color=color,
                fill_opacity=0.75,
                weight=2,
                popup=folium.Popup(popup_html, max_width=340),
                tooltip=f"{mun} · {n_p} proy. · {valor}",
            ).add_to(m)

            # Etiqueta con número de proyectos encima del marcador
            folium.Marker(
                location=[lat, lng],
                icon=folium.DivIcon(
                    html=f'<div style="font-family:Arial;font-size:11px;font-weight:800;'
                         f'color:#fff;text-align:center;text-shadow:0 1px 2px rgba(0,0,0,.5);'
                         f'margin-top:-6px">{n_p}</div>',
                    icon_size=(30, 20),
                    icon_anchor=(15, 10),
                )
            ).add_to(m)

        # ── Plugins adicionales ──────────────────────────────────────────────
        from folium.plugins import Fullscreen, MiniMap
        Fullscreen(
            position="topright",
            title="Pantalla completa",
            force_separate_button=True,
        ).add_to(m)
        MiniMap(toggle_display=True, position="bottomright").add_to(m)

        # Control de capas base
        folium.TileLayer("CartoDB dark_matter", name="Oscuro",  control=True).add_to(m)
        folium.TileLayer("OpenStreetMap",       name="Calles",   control=True).add_to(m)
        folium.LayerControl(position="topright", collapsed=True).add_to(m)

        # Leyenda
        legend_html = """
<div style="position:fixed;bottom:20px;left:20px;z-index:999;
     background:white;border:1px solid #ccc;border-radius:8px;padding:12px 16px;
     font-family:Arial;font-size:12px;box-shadow:0 2px 8px rgba(0,0,0,.1)">
  <b style="font-size:13px;color:#003d6c">Proyectos por municipio</b><br><br>
  <span style="color:#47b1d5">●</span> Sin proyectos &nbsp;
  <span style="color:#17743d">●</span> 1–2 &nbsp;
  <span style="color:#1754ab">●</span> 3–5 &nbsp;
  <span style="color:#003d6c">●</span> 6+<br>
  <span style="font-size:11px;color:#888">El número indica cantidad de proyectos por municipio</span>
</div>"""
        m.get_root().html.add_child(folium.Element(legend_html))

        # Renderizar con captura de clics
        map_data = st_folium(
            m,
            use_container_width=True,
            height=580,
            returned_objects=["last_object_clicked_tooltip"],
        )

        # Si hicieron clic en un municipio, mostrar su detalle debajo del mapa
        clicked = map_data.get("last_object_clicked_tooltip") if map_data else None
        if clicked:
            mun_click = clicked.split(" · ")[0].strip() if " · " in clicked else clicked.strip()
            if mun_click in mun_data_render:
                d           = mun_data_render[mun_click]
                _n_p        = len(d["proyectos"])
                _n_c        = len(d["convocatorias"])
                _val        = fmt_money(d["valor"])
                _ben        = f"{d['beneficiarios']:,}"
                _proy_html  = "".join(
                    f'<div style="font-size:.8rem;color:#333;padding:4px 0;'
                    f'border-bottom:1px solid #f0f0f0">{p}</div>'
                    for p in d["proyectos"]
                )
                _kpis = (
                    kpi("Proyectos",     str(_n_p), "formulados",  border_color="#1754ab", flex="1") +
                    kpi("Convocatorias", str(_n_c), "vinculadas",  border_color="#17743d", flex="1") +
                    kpi("Valor total",   _val,      "en proyectos",border_color="#cf7000", flex="1") +
                    kpi("Beneficiarios", _ben,      "en municipio",border_color="#47b1d5", flex="1")
                )
                st.markdown(
                    f'<div style="background:#f8fbff;border:1px solid #cce0f5;'
                    f'border-left:5px solid #1754ab;border-radius:10px;'
                    f'padding:20px 26px;margin:12px 0">'
                    f'<div style="font-family:\'DM Serif Display\',serif;font-size:1.3rem;'
                    f'color:#003d6c;margin-bottom:14px">{mun_click}</div>'
                    f'<div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:16px">{_kpis}</div>'
                    f'<div style="font-size:.82rem;font-weight:600;color:#003d6c;margin-bottom:6px">'
                    f'Proyectos:</div>{_proy_html}</div>',
                    unsafe_allow_html=True,
                )

        # Tabla resumen de municipios
        if mun_data_render:
            st.markdown(sec_title("Detalle por municipio"), unsafe_allow_html=True)
            mun_rows = []
            for mun, data in sorted(mun_data_render.items(), key=lambda x: -len(x[1]["proyectos"])):
                if mun in SUCRE_COORDS:
                    mun_rows.append({
                        "Municipio":     mun,
                        "Proyectos":     len(data["proyectos"]),
                        "Convocatorias": len(data["convocatorias"]),
                        "Valor total":   data["valor"],
                        "Beneficiarios": data["beneficiarios"],
                        "Nombres proyectos": " · ".join(data["proyectos"]),
                    })
            if mun_rows:
                df_mun = pd.DataFrame(mun_rows)
                st.dataframe(df_mun.reset_index(drop=True),
                    use_container_width=True, height=min(420, 60+len(mun_rows)*42),
                    hide_index=True,
                    column_config={
                        "Municipio":         st.column_config.TextColumn(width=160),
                        "Proyectos":         st.column_config.NumberColumn(width=90),
                        "Convocatorias":     st.column_config.NumberColumn(width=110),
                        "Valor total":       st.column_config.NumberColumn("Valor $", format="$ %d"),
                        "Beneficiarios":     st.column_config.NumberColumn(width=110),
                        "Nombres proyectos": st.column_config.TextColumn(width=380),
                    })

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 · ASISTENTE IA (Gemini)
# ─────────────────────────────────────────────────────────────────────────────
with tab4:
    st.markdown(sec_title("Asistente IA",
        "Consulta los datos de convocatorias y proyectos en lenguaje natural · Powered by Gemini"),
        unsafe_allow_html=True)

    @st.cache_data(ttl=300, show_spinner=False)
    def _build_context(ch, ph):
        def _to_csv(df, name, max_rows=300):
            if df is None or df.empty: return f"[{name}: sin datos]\n"
            cols = [c for c in df.columns if c not in ("id","convocatoria_id","proyecto_id")]
            return f"=== {name.upper()} ({len(df)} registros) ===\n{df[cols].head(max_rows).to_csv(index=False)}\n"
        ctx  = _to_csv(df_conv, "Convocatorias")
        ctx += _to_csv(df_proy, "Proyectos")
        ctx += _to_csv(df_rel,  "Relaciones (Convocatoria–Proyecto)")
        ctx += _to_csv(df_ind,  "Indicadores MGA")
        return ctx

    _ch = str(len(df_conv)) + str(int(df_conv["id"].max()) if not df_conv.empty else 0)
    _ph = str(len(df_proy)) + str(int(df_proy["id"].max()) if not df_proy.empty else 0)
    data_context = _build_context(_ch, _ph)

    SYSTEM_PROMPT = f"""Eres un asistente de análisis de datos especializado en convocatorias \
y proyectos de la Secretaría de Planeación (SDP) de Bogotá.

Tienes acceso a la base de datos completa cargada desde Supabase:

{data_context}

INSTRUCCIONES:
- Responde siempre en español, de forma clara y concisa.
- Extrae la respuesta directamente de los datos. No inventes cifras.
- Si la respuesta incluye una tabla, usa Markdown con columnas alineadas.
- Puedes calcular: sumas, promedios, porcentajes, rankings, filtros por sector/estado/dependencia.
- Respuestas máximo 400 palabras salvo que el usuario pida más detalle.
"""

    def _call_gemini(messages):
        import urllib.request, json as _json
        url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
               f"{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}")
        contents = [
            {"role":"user",  "parts":[{"text":f"[CONTEXTO]\n{SYSTEM_PROMPT}"}]},
            {"role":"model", "parts":[{"text":"Entendido. Listo para responder sobre la SDP."}]},
        ]
        for m in messages:
            contents.append({"role": m["role"], "parts":[{"text": m["content"]}]})
        body = _json.dumps({
            "contents": contents,
            "generationConfig": {"temperature":0.3,"maxOutputTokens":1500,"topP":0.9},
        }).encode()
        req = urllib.request.Request(url, data=body, headers={"Content-Type":"application/json"})
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                data = _json.loads(resp.read())
            return data["candidates"][0]["content"]["parts"][0]["text"]
        except urllib.error.HTTPError as e:
            err = e.read().decode()
            return f"Error HTTP {e.code}: {err[:300]}"
        except Exception as ex:
            return f"Error al conectar con Gemini: {ex}"

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    suggestions = [
        "¿Cuántas convocatorias hay vigentes?",
        "¿Cuál es el sector con más convocatorias?",
        "Lista los 5 proyectos de mayor valor",
        "¿Cuánto suman los proyectos por dependencia?",
        "¿Qué convocatorias no tienen proyectos?",
        "¿Cuál es el promedio de cobertura financiera?",
        "¿Cuáles son los indicadores MGA más usados?",
        "Resumen general de la base de datos",
    ]

    st.markdown('<div style="font-size:.8rem;color:#666;font-weight:600;margin-bottom:10px">'
                'Preguntas sugeridas:</div>', unsafe_allow_html=True)
    cols_s = st.columns(4)
    for idx, sug in enumerate(suggestions):
        if cols_s[idx%4].button(sug, key=f"sug_{idx}", use_container_width=True):
            st.session_state.chat_history.append({"role":"user","content":sug})
            with st.spinner("Consultando Gemini..."):
                resp = _call_gemini(st.session_state.chat_history)
            st.session_state.chat_history.append({"role":"model","content":resp})
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    if st.session_state.chat_history:
        chat_html = '<div class="chat-scroll">'
        for msg in st.session_state.chat_history:
            if msg["role"] == "user":
                chat_html += (f'<div style="display:flex;justify-content:flex-end;margin:8px 0">'
                              f'<div class="chat-user">{msg["content"]}</div></div>')
            else:
                chat_html += (f'<div style="display:flex;justify-content:flex-start;margin:8px 0">'
                              f'<div class="chat-ai">'
                              f'<div class="chat-ai-label">Gemini · Asistente IA</div>'
                              f'<div style="white-space:pre-wrap">{msg["content"]}</div>'
                              f'</div></div>')
        chat_html += '</div>'
        st.markdown(chat_html, unsafe_allow_html=True)
    else:
        st.markdown(
            '<div style="background:#f0f8fb;border:1px dashed #47b1d5;border-radius:10px;'
            'padding:30px;text-align:center;color:#666;margin-bottom:16px">'
            'Escribe una pregunta o usa las sugerencias para empezar</div>',
            unsafe_allow_html=True)

    col_inp, col_btn, col_clr = st.columns([6,1,1])
    with col_inp:
        user_input = st.text_input("Pregunta", placeholder="Ej: ¿Cuál es el estado con más convocatorias?",
                                   key="chat_input", label_visibility="collapsed")
    with col_btn:
        send = st.button("Enviar", type="primary", use_container_width=True)
    with col_clr:
        if st.button("Limpiar", use_container_width=True):
            st.session_state.chat_history = []; st.rerun()

    if send and user_input.strip():
        st.session_state.chat_history.append({"role":"user","content":user_input.strip()})
        with st.spinner("Consultando Gemini..."):
            resp = _call_gemini(st.session_state.chat_history)
        st.session_state.chat_history.append({"role":"model","content":resp})
        st.rerun()

    st.markdown(
        f'<div style="font-size:.72rem;color:#aaa;margin-top:8px;text-align:right">'
        f'Contexto: {len(data_context):,} caracteres · '
        f'{len(df_conv)} convocatorias · {len(df_proy)} proyectos · {len(df_ind)} indicadores </div>',
        unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 · EXPORTAR
# ─────────────────────────────────────────────────────────────────────────────
with tab5:
    st.markdown(sec_title("Exportar Reporte Maestro",
        "Generación de sábana de datos consolidada (.xlsx)"), unsafe_allow_html=True)

    opt = st.radio("Alcance",
                   ["Exportar universo completo (sin filtros)",
                    "Exportar selección actual (datos filtrados)"],
                   horizontal=True)
    ec = df_c if "filtrados" in opt else df_conv
    ep = df_p if "filtrados" in opt else df_proy
    er = df_r if "filtrados" in opt else df_rel
    ei = df_i if "filtrados" in opt else df_ind

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"""
<div style="display:flex;gap:14px;margin-bottom:24px;flex-wrap:wrap;">
    {kpi("Convocatorias", ec["id"].nunique() if not ec.empty else 0, "a exportar", border_color="#1754ab")}
    {kpi("Proyectos",     ep["id"].nunique() if not ep.empty else 0, "a exportar", border_color="#1754ab")}
    {kpi("Relaciones",    len(er) if not er.empty else 0,            "a exportar", border_color="#1754ab")}
    {kpi("Indicadores",   len(ei) if not ei.empty else 0,            "a exportar", border_color="#1754ab")}
</div>""", unsafe_allow_html=True)

    if st.button("Generar reporte de Excel", type="primary"):
        with st.spinner("Construyendo matriz Excel..."):
            H_FILL = PatternFill("solid", fgColor="003d6c")
            H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            C_FONT = Font(name="Arial", size=9)
            WHITE  = PatternFill("solid", fgColor="FFFFFF")
            THIN   = Border(left=Side(style="thin",color="CCCCCC"),right=Side(style="thin",color="CCCCCC"),
                            top=Side(style="thin",color="CCCCCC"),bottom=Side(style="thin",color="CCCCCC"))
            wb = Workbook(); wb.remove(wb.active)
            sheets_def = [
                ("Convocatorias", ec, [
                    "Convocatoria","Estado","Fecha apertura","Fecha cierre","Monto",
                    "Sectores","Segmentos","Ubicación","Dependencias","Aliados",
                    "N° proyectos","Valor proyectos","Cobertura (%)","Contacto",
                    "Qué ofrece","Quiénes participan","Público priorizado"]),
                ("Proyectos", ep, [
                    "Proyecto","BPIN","Valor","Contrapartida","Dependencia","Responsable",
                    "Municipios","Total beneficiarios","Tipos beneficiarios",
                    "N° indicadores MGA","Indicadores MGA"]),
                ("Relaciones", er, [
                    "Convocatoria","Estado convocatoria","Sectores","Proyecto","BPIN","Valor",
                    "Dependencia","Responsable","Fecha apertura","Fecha cierre","Cobertura (%)"]),
                ("Indicadores MGA", ei, [
                    "Proyecto","codigo","nombre","vigencia","meta_proyecto","meta_cuatrienio",
                    "m2024","m2025","m2026","m2027","responsable_mga"]),
            ]
            COL_W = {"Convocatoria":42,"Proyecto":42,"nombre":42,"Qué ofrece":50,
                     "Quiénes participan":40,"Sectores":22,"Segmentos":20,
                     "Municipios":22,"Tipos beneficiarios":30,"Indicadores MGA":40}
            for sname, df, cols in sheets_def:
                if df is None or df.empty: continue
                cols = [c for c in cols if c in df.columns]
                ws   = wb.create_sheet(sname)
                ws.sheet_view.showGridLines = False
                for ci, col in enumerate(cols,1):
                    c = ws.cell(row=1,column=ci,value=col)
                    c.font=H_FONT; c.fill=H_FILL
                    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
                    c.border=THIN
                ws.row_dimensions[1].height=28
                for ri,(_,row) in enumerate(df[cols].iterrows(),2):
                    for ci,col in enumerate(cols,1):
                        val=row[col]
                        if pd.isna(val): val=""
                        c=ws.cell(row=ri,column=ci,value=val)
                        c.font=C_FONT; c.fill=WHITE; c.border=THIN
                        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
                    ws.row_dimensions[ri].height=30
                for ci,col in enumerate(cols,1):
                    ws.column_dimensions[get_column_letter(ci)].width=COL_W.get(col,15)
                ws.freeze_panes="A2"
                tname="T_"+re.sub(r"[^A-Za-z0-9]","_",sname)
                tbl=Table(displayName=tname,ref=f"A1:{get_column_letter(len(cols))}{1+len(df)}")
                tbl.tableStyleInfo=TableStyleInfo(name="TableStyleMedium7",showRowStripes=False)
                ws.add_table(tbl)
            buf=io.BytesIO(); wb.save(buf)
        st.success("La matriz ha sido generada con éxito.")
        st.download_button("Descargar Reporte_SDP.xlsx", data=buf.getvalue(),
            file_name="Reporte_SDP.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown('<hr style="border-color:#e0e0e0;margin-top:40px;margin-bottom:10px">',
            unsafe_allow_html=True)
st.markdown('<div style="text-align:center;padding:10px;font-size:.85rem;color:#888">'
            'Secretaría de Planeación · Seguimiento de Convocatorias y Proyectos'
            '</div>', unsafe_allow_html=True)
