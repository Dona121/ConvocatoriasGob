import io
import re
import math
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Gestión de Convocatorias SDP",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Colores del Diseño (Aclarado) ──────────────────────────────────────────────────
# Un verde vibrante y visible sobre fondo claro.
GREEN_PRIMARY = "#00C853"
GREEN_DARK    = "#008138"  # Para hover y textos pequeños.
BG_LIGHT_SIDE = "#F8FAFC" # Gris muy suave para sidebar.
BG_LIGHT      = "#FFFFFF" # Blanco puro para el fondo principal.
BORDER_COLOR  = "#E2E8F0" # Color de borde muy tenue.
TEXT_MAIN     = "#1A1D29" # Gris muy oscuro para texto principal.
TEXT_SUB      = "#64748B" # Gris medio para subtítulos.

# Paleta aclarada para gráficos (verdes frescos).
GREENS_PALETTE = ["#00E676", "#66FFB3", "#81F7D7", "#B9F6CA", "#DCEDC8",
                  "#44D5AB", "#76DBCB", "#A2E4DA", "#B5EAE2", "#D2F2EF"]

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&display=swap');

html, body, [class*="css"] {{ 
    font-family: 'Plus Jakarta Sans', sans-serif; 
    color: {TEXT_MAIN};
    background-color: {BG_LIGHT};
}}

/* Sidebar - Fondo claro con borde sutil */
section[data-testid="stSidebar"] > div:first-child {{
    background-color: {BG_LIGHT_SIDE} !important;
    border-right: 1px solid {BORDER_COLOR};
}}

/* Tabs - Diseño Minimalista y Claro */
.stTabs [data-baseweb="tab-list"] {{
    border-bottom: 2px solid {BORDER_COLOR};
    gap: 8px;
    background-color: {BG_LIGHT};
}}
.stTabs [data-baseweb="tab"] {{
    font-weight: 600;
    font-size: 0.9rem;
    color: {TEXT_SUB};
    background-color: transparent;
    border-radius: 6px 6px 0 0;
    padding: 10px 24px;
}}
.stTabs [aria-selected="true"] {{
    color: {GREEN_PRIMARY} !important;
    border-bottom: 3px solid {GREEN_PRIMARY} !important;
}}

/* Botones - Verdes Brillantes */
/* Download button */
.stDownloadButton > button {{
    background-color: {GREEN_PRIMARY} !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    padding: 10px 24px !important;
    font-size: 0.9rem;
}}
.stDownloadButton > button:hover {{
    background-color: {GREEN_DARK} !important;
    box-shadow: 0 2px 6px rgba(0, 129, 56, 0.15);
}}

/* Primary button (radio group) */
.stButton > button[kind="primary"] {{
    background-color: {GREEN_PRIMARY} !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
}}

</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
TABLE_NAME = "SeguimientoConvocatorias"
INVALID_SECTORS = {"", "none", "nan", "n/d", "no especifica",
                   "verificar", "formuladores-evaluadores", "varios"}

COLS_REPORT = [
    "ID", "NOMBRE DE LA CONVOCATORIA", "SEGMENTO",
    "FECHA DE APERTURA", "FECHA DE CIERRE", "DÍAS DISPONIBLES",
    "ESTADO", "MONTO POR PROYECTO", "OBJETIVO",
    "CONTACTO", "QUIENES PUEDEN PARTICIPAR", "FUENTES",
]
COL_WIDTHS = {
    "ID": 6, "NOMBRE DE LA CONVOCATORIA": 38, "SEGMENTO": 22,
    "FECHA DE APERTURA": 18, "FECHA DE CIERRE": 18, "DÍAS DISPONIBLES": 12,
    "ESTADO": 10, "MONTO POR PROYECTO": 16, "OBJETIVO": 50,
    "CONTACTO": 35, "QUIENES PUEDEN PARTICIPAR": 30, "FUENTES": 20,
}


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
def _read_named_table(file_bytes: bytes, table_name: str) -> pd.DataFrame:
    """
    Lee una tabla de Excel por nombre usando openpyxl.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if tbl.name == table_name:
                data = list(ws[tbl.ref])
                headers = [cell.value for cell in data[0]]
                rows = [[cell.value for cell in row] for row in data[1:]]
                return pd.DataFrame(rows, columns=headers)
    raise ValueError(
        f"No se encontró la tabla '{table_name}' en el archivo. "
        f"Verifica que el Excel contenga una tabla con ese nombre exacto."
    )


@st.cache_data(show_spinner=False)
def load_data(file_bytes: bytes):
    """
    Lee SeguimientoConvocatorias, normaliza SECTOR y explota las categorías.
    """
    df = _read_named_table(file_bytes, TABLE_NAME)

    # Normalizar SECTOR
    df["SECTOR"] = df["SECTOR"].astype(str).str.strip()
    df = df[~df["SECTOR"].str.lower().isin(INVALID_SECTORS)].copy()
    df = df.reset_index(drop=True)

    base = df.copy()

    # Explotar la columna SECTOR
    exploded = base.copy()
    exploded["SECTOR"] = exploded["SECTOR"].str.split(" - ")
    exploded = exploded.explode("SECTOR")
    exploded["SECTOR"] = exploded["SECTOR"].str.strip()
    exploded = exploded[~exploded["SECTOR"].str.lower().isin(INVALID_SECTORS)]
    exploded = exploded.reset_index(drop=True)

    return base, exploded


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT BUILDER (FORMATO CLARO)
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(exploded: pd.DataFrame) -> bytes:
    """
    Genera el archivo Excel final con pestañas por sector y tablas con formato.
    """
    H_FILL = PatternFill("solid", fgColor="DCEDC8") # Verde muy claro para cabeceras.
    H_FONT = Font(bold=True, color="004D1B", name="Arial", size=10) # Texto verde oscuro.
    T_FONT = Font(bold=True, color="008138", name="Arial", size=13) # Título verde vibrante.
    C_FONT = Font(name="Arial", size=9)
    WHITE  = PatternFill("solid", fgColor="FFFFFF")
    GRAY_SUB = Font(name="Arial", size=9, color="757575", italic=True)
    
    THIN   = Border(
        left=Side(style="thin", color="EEEEEE"), # Borde muy suave.
        right=Side(style="thin", color="EEEEEE"),
        top=Side(style="thin", color="EEEEEE"),
        bottom=Side(style="thin", color="EEEEEE"),
    )

    wb = Workbook()
    wb.remove(wb.active) # Eliminar hoja por defecto.
    sectores = sorted(exploded["SECTOR"].unique())
    available_cols = [c for c in COLS_REPORT if c in exploded.columns]

    # ── Hoja índice ──
    wi = wb.create_sheet("Índice")
    wi.sheet_view.showGridLines = False
    wi["A1"] = "Convocatorias por Sector"
    wi["A1"].font = Font(bold=True, color="008138", name="Arial", size=15)
    
    # Cabeceras
    for ci, label in enumerate(["SECTOR", "N° CONVOCATORIAS"], 1):
        c = wi.cell(row=3, column=ci, value=label)
        c.font = Font(bold=True, color="004D1B", name="Arial", size=10)
        c.fill = H_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN
    
    # Contenido
    for i, s in enumerate(sectores, 4):
        n = exploded[exploded["SECTOR"] == s]["ID"].nunique()
        for ci, val in enumerate([s, n], 1):
            c = wi.cell(row=i, column=ci, value=val)
            c.font = C_FONT; c.fill = WHITE; c.border = THIN
            c.alignment = Alignment(
                horizontal="center" if ci == 2 else "left", vertical="center"
            )
    wi.column_dimensions["A"].width = 32
    wi.column_dimensions["B"].width = 20
    tbl_i = Table(displayName="Indice", ref=f"A3:B{3 + len(sectores)}")
    # Usar estilo claro.
    tbl_i.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    wi.add_table(tbl_i)

    # ── Una hoja por sector ──
    for sector in sectores:
        sname = sector[:31].replace("/", "-").replace("\\", "-").replace(":", "")
        ws = wb.create_sheet(sname)
        ws.sheet_view.showGridLines = False

        # Título de la hoja.
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(available_cols))
        tc = ws.cell(row=1, column=1, value=f"Sector: {sector}")
        tc.font = T_FONT
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        subset = (exploded[exploded["SECTOR"] == sector][available_cols]
                  .reset_index(drop=True))
        
        # Subtítulo (conteo).
        nc = ws.cell(row=2, column=1, value=f"{len(subset)} convocatoria(s) registradas")
        nc.font = GRAY_SUB
        ws.row_dimensions[2].height = 14

        # Cabeceras de tabla.
        for ci, col in enumerate(available_cols, 1):
            c = ws.cell(row=3, column=ci, value=col)
            c.font = H_FONT; c.fill = H_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN
        ws.row_dimensions[3].height = 30

        # Contenido de tabla.
        for ri, (_, row) in enumerate(subset.iterrows(), 4):
            for ci, col in enumerate(available_cols, 1):
                val = row[col]
                if pd.isna(val): val = ""
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = C_FONT; c.fill = WHITE; c.border = THIN
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Fila blanca, alta para wrap_text.
            ws.row_dimensions[ri].height = 45

        # Ajuste de ancho de columna.
        for ci, col in enumerate(available_cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)

        # Congelar paneles y añadir formato de tabla Excel.
        ws.freeze_panes = "A4"
        last_col = get_column_letter(len(available_cols))
        tname = "T_" + re.sub(r"[^A-Za-z0-9_]", "_", sector)[:28]
        tbl = Table(displayName=tname,
                    ref=f"A3:{last_col}{3 + len(subset)}")
        # Estilo claro.
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(tbl)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# HTML CHART HELPERS (Aclarados)
# ══════════════════════════════════════════════════════════════════════════════
def bar_chart(data: pd.Series, title: str, max_bars: int = 25) -> str:
    """
    Gráfico de barras horizontal HTML/CSS simplificado y claro.
    """
    data = data.sort_values(ascending=False).head(max_bars)
    max_val = data.max() or 1
    rows = ""
    for i, (label, val) in enumerate(data.items()):
        pct = round((val / max_val) * 100, 1)
        # Verde brillante, degradando sutilmente.
        intensity = max(0.65, 1 - i * 0.015)
        color = f'rgba(0, 200, 83, {intensity:.2f})' 
        
        rows += (
            f'<div style="display:flex;align-items:center;margin-bottom:8px;gap:12px">'
            f'<div style="width:175px;font-size:0.8rem;color:{TEXT_SUB};text-align:right;'
            f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;flex-shrink:0" '
            f'title="{label}">{label}</div>'
            f'<div style="flex:1;background:#F1F5F9;border-radius:4px;height:22px;position:relative">'
            f'<div style="width:{pct}%;background:{color};height:100%;border-radius:4px"></div>'
            f'<span style="position:absolute;right:8px;top:4px;font-size:0.75rem;'
            f'font-weight:700;color:white;text-shadow:0 1px 2px rgba(0,0,0,0.2)">{val}</span>'
            f'</div></div>'
        )
    return (
        f'<div style="background:{BG_LIGHT};border:1px solid {BORDER_COLOR};border-radius:10px;padding:22px 24px 18px;box-shadow: 0 1px 3px rgba(0,0,0,0.02);">'
        f'<div style="font-size:1.1rem;font-weight:600;color:{TEXT_MAIN};'
        f'margin-bottom:14px;padding-bottom:10px;border-bottom:2px solid {BORDER_COLOR}">{title}</div>'
        f'{rows}</div>'
    )


def donut_chart(data: pd.Series, title: str, top_n: int = 8) -> str:
    """
    Gráfico de rosca SVG simplificado y claro.
    """
    total = data.sum()
    if total == 0:
        return ""
    top = data.sort_values(ascending=False).head(top_n)
    
    # Config SVG
    cx, cy, r, ir = 75, 75, 60, 36
    angle = -90.0
    paths = ""
    
    for i, (_, val) in enumerate(top.items()):
        sweep = (val / total) * 360
        end = angle + sweep
        
        a1r, a2r = math.radians(angle), math.radians(end)
        x1, y1 = cx + r * math.cos(a1r), cy + r * math.sin(a1r)
        x2, y2 = cx + r * math.cos(a2r), cy + r * math.sin(a2r)
        ix1, iy1 = cx + ir * math.cos(a2r), cy + ir * math.sin(a2r)
        ix2, iy2 = cx + ir * math.cos(a1r), cy + ir * math.sin(a1r)
        large = 1 if sweep > 180 else 0
        
        color = GREENS_PALETTE[i % len(GREENS_PALETTE)]
        
        paths += (
            f'<path d="M{x1:.1f},{y1:.1f} A{r},{r} 0 {large},1 {x2:.1f},{y2:.1f} '
            f'L{ix1:.1f},{iy1:.1f} A{ir},{ir} 0 {large},0 {ix2:.1f},{iy2:.1f} Z" '
            f'fill="{color}" stroke="white" stroke-width="2"/>'
        )
        angle = end

    # Leyenda
    legend = ""
    for i, (label, val) in enumerate(top.items()):
        pct = round(val / total * 100, 1)
        color = GREENS_PALETTE[i % len(GREENS_PALETTE)]
        legend += (
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">'
            f'<div style="width:10px;height:10px;border-radius:50%;background:{color};flex-shrink:0"></div>'
            f'<div style="font-size:0.8rem;color:{TEXT_SUB};flex:1;white-space:nowrap;'
            f'overflow:hidden;text-overflow:ellipsis" title="{label}">{label}</div>'
            f'<div style="font-size:0.8rem;font-weight:700;color:{TEXT_MAIN}">{pct}%</div>'
            f'</div>'
        )

    # SVG central (total)
    svg = (
        f'<svg width="150" height="150" viewBox="0 0 150 150">{paths}'
        f'<text x="{cx}" y="{cy + 5}" text-anchor="middle" font-size="20" '
        f'fill="{TEXT_MAIN}" font-weight="700">{total}</text>'
        f'<text x="{cx}" y="{cy + 19}" text-anchor="middle" font-size="9" '
        f'fill="{TEXT_SUB}">total</text></svg>'
    )

    return (
        f'<div style="background:{BG_LIGHT};border:1px solid {BORDER_COLOR};border-radius:10px;padding:22px 24px 18px;box-shadow: 0 1px 3px rgba(0,0,0,0.02);">'
        f'<div style="font-size:1.1rem;font-weight:600;color:{TEXT_MAIN};'
        f'margin-bottom:14px;padding-bottom:10px;border-bottom:2px solid {BORDER_COLOR}">{title}</div>'
        f'<div style="display:flex;gap:24px;align-items:center">'
        f'<div style="flex-shrink:0">{svg}</div>'
        f'<div style="flex:1;overflow:hidden">{legend}</div>'
        f'</div></div>'
    )


def metric_card(label: str, value, sub: str) -> str:
    """
    Tarjeta KPI clara con indicador lateral verde.
    """
    return (
        f'<div style="background:{BG_LIGHT};border:1px solid {BORDER_COLOR};border-left:4px solid {GREEN_PRIMARY};'
        f'border-radius:8px;padding:20px 22px;margin-bottom:8px;box-shadow: 0 1px 3px rgba(0,0,0,0.02);">'
        f'<div style="font-size:0.75rem;letter-spacing:0.08em;text-transform:uppercase;'
        f'color:{TEXT_SUB};font-weight:600;margin-bottom:4px">{label}</div>'
        f'<div style="font-size:2.2rem;'
        f'color:{TEXT_MAIN};line-height:1;font-weight:700">{value}</div>'
        f'<div style="font-size:0.82rem;color:{TEXT_SUB};margin-top:4px">{sub}</div>'
        f'</div>'
    )


def section_title(text: str, sub: str = "") -> str:
    """
    Título de sección con línea divisoria clara.
    """
    sub_html = (
        f'<div style="font-size:0.88rem;color:{TEXT_SUB};margin-bottom:18px">{sub}</div>'
        if sub else ""
    )
    return (
        f'<div style="font-size:1.4rem;font-weight:600;color:{TEXT_MAIN};'
        f'margin:32px 0 8px;padding-bottom:10px;border-bottom:2px solid {BORDER_COLOR}">{text}</div>'
        f'{sub_html}'
    )


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(
        f'<div style="padding:18px 0 20px">'
        f'<div style="font-size:1.4rem;'
        f'color:{TEXT_MAIN};line-height:1.25;font-weight:700">Convocatorias</div>'
        f'<div style="color:{TEXT_SUB};font-size:0.85rem;font-weight:300;margin-top:4px">'
        f'SDP · Reporte Interactivo</div></div>'
        f'<hr style="border-color:{BORDER_COLOR};margin-bottom:20px">',
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader(
        "Cargar archivo Excel",
        type=["xlsx"],
        help=(
            "Debe contener la tabla 'SeguimientoConvocatorias' "
            "con la columna SECTOR ya estandarizada."
        ),
    )

    if uploaded:
        st.markdown(
            f'<hr style="border-color:{BORDER_COLOR};margin:20px 0 14px">'
            f'<div style="font-size:0.7rem;letter-spacing:0.1em;text-transform:uppercase;'
            f'color:{TEXT_SUB};font-weight:600;margin-bottom:12px">Configuración del Reporte</div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# LANDING (Fondo Claro con Gradiente)
# ══════════════════════════════════════════════════════════════════════════════
if not uploaded:
    # Gradiente de blanco a verde muy tenue.
    st.markdown(
        f'<div style="background:linear-gradient(135deg,{BG_LIGHT} 0%,{GREENS_PALETTE[4]} 100%);'
        f'border:1px solid {BORDER_COLOR};'
        f'border-radius:12px;padding:48px 42px 42px;margin-bottom:32px;text-align:center">'
        f'<div style="font-size:2.2rem;'
        f'color:{TEXT_MAIN};margin:0 0 10px;line-height:1.2;font-weight:700">Reporte de Convocatorias SDP</div>'
        f'<div style="color:{TEXT_SUB};font-size:0.95rem;font-weight:400;max-width:600px;margin:0 auto">'
        f'Por favor, carga el archivo Excel que contiene la tabla SeguimientoConvocatorias '
        f'para habilitar el dashboard analítico y la generación de reportes automáticos.</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    
    c1, c2, c3 = st.columns(3)
    for col, lbl, sub in [
        (c1, "Dashboard Visual",     "Análisis gráfico por sector, segmento y estado."),
        (c2, "Explorador de Datos", "Tabla interactiva filtrable con detalle por sector."),
        (c3, "Reporte Excel Automatizado", "Generación de documento estructurado con hojas por sector."),
    ]:
        # Tarjeta blanca clara con indicador superior verde.
        col.markdown(
            f'<div style="background:{BG_LIGHT};border:1px solid {BORDER_COLOR};border-top:4px solid {GREEN_PRIMARY};'
            f'border-radius:8px;padding:24px;text-align:center;box-shadow: 0 1px 3px rgba(0,0,0,0.02);">'
            f'<div style="font-size:0.8rem;letter-spacing:0.08em;text-transform:uppercase;'
            f'color:{TEXT_MAIN};font-weight:700;margin-bottom:8px">{lbl}</div>'
            f'<div style="font-size:0.85rem;color:{TEXT_SUB};line-height:1.5">{sub}</div></div>',
            unsafe_allow_html=True,
        )
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════
with st.spinner("Procesando tabla SeguimientoConvocatorias…"):
    file_bytes = uploaded.read()
    try:
        base_df, exploded_df = load_data(file_bytes)
    except ValueError as e:
        st.error(str(e))
        st.stop()
    except Exception as e:
        st.error(f"Error inesperado al leer el archivo: {e}")
        st.stop()

if base_df.empty:
    st.warning("La tabla no contiene registros válidos en la columna SECTOR.")
    st.stop()

# Listas para filtros.
sectores_all  = sorted(exploded_df["SECTOR"].unique())
segmentos_all = sorted(base_df["SEGMENTO"].dropna().unique()) if "SEGMENTO" in base_df.columns else []
estados_all   = sorted(base_df["ESTADO"].dropna().unique())   if "ESTADO"   in base_df.columns else []

# ── Filtros en sidebar ─────────────────────────────────────────────────────────
with st.sidebar:
    sel_sectores  = st.multiselect("Filtrar por Sector", sectores_all,  placeholder="Todos")
    sel_segmentos = st.multiselect("Filtrar por Segmento", segmentos_all, placeholder="Todos") if segmentos_all else []
    sel_estados   = st.multiselect("Filtrar por Estado",   estados_all,   placeholder="Todos") if estados_all   else []

# Aplicar filtros.
exp_f  = exploded_df.copy()
base_f = base_df.copy()

if sel_sectores:
    exp_f  = exp_f[exp_f["SECTOR"].isin(sel_sectores)]
    base_f = base_f[base_f["ID"].isin(exp_f["ID"])] # Asegurar coincidencia de ID.
if sel_segmentos:
    base_f = base_f[base_f["SEGMENTO"].isin(sel_segmentos)]
    exp_f  = exp_f[exp_f["ID"].isin(base_f["ID"])]
if sel_estados:
    base_f = base_f[base_f["ESTADO"].isin(sel_estados)]
    exp_f  = exp_f[exp_f["ID"].isin(base_f["ID"])]


# ══════════════════════════════════════════════════════════════════════════════
# HERO + KPIs (Aclarado)
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(
    f'<div style="background:{BG_LIGHT_SIDE};'
    f'border-radius:12px;padding:36px 40px;margin-bottom:30px;'
    f'border:1px solid {BORDER_COLOR};box-shadow: 0 1px 3px rgba(0,0,0,0.02)">'
    f'<div style="font-size:1.8rem;font-weight:700;'
    f'color:{TEXT_MAIN};margin:0 0 6px;line-height:1.2">Análisis de Convocatorias Registradas</div>'
    f'<div style="color:{TEXT_SUB};font-size:0.92rem;font-weight:300">'
    f'Documento: {uploaded.name} · {len(base_df)} registros · '
    f'{len(sectores_all)} sectores temáticos</div></div>',
    unsafe_allow_html=True,
)

# KPIs siempre sobre base_f (antes del explode) — conteo único por ID.
n_conv     = base_f['ID'].nunique() if 'ID' in base_f.columns else len(base_f)
n_vigentes = (
    len(base_f[base_f["ESTADO"].astype(str).str.upper().str.contains("VIGENTE", na=False)])
    if "ESTADO" in base_f.columns else 0
)
pct_vig    = round(n_vigentes / max(n_conv, 1) * 100)
n_sectores = exp_f["SECTOR"].nunique()  # sectores atómicos activos tras filtro.
n_segmentos = base_f["SEGMENTO"].nunique() if "SEGMENTO" in base_f.columns else 0

k1, k2, k3, k4 = st.columns(4)
k1.markdown(metric_card("Convocatorias Totales", n_conv,     "Registros únicos evaluados"), unsafe_allow_html=True)
k2.markdown(metric_card("Registros Vigentes",      n_vigentes, f"{pct_vig}% del total actual"), unsafe_allow_html=True)
k3.markdown(metric_card("Categorías Temáticas",      n_sectores, "Sectores activos en filtros"), unsafe_allow_html=True)
k4.markdown(metric_card("Tipos de Segmento",     n_segmentos,"Segmentos distintos"), unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
tab1, tab2, tab3 = st.tabs(["Dashboard General", "Explorador de Datos", "Generar Reporte Excel"])


# ─── TAB 1: DASHBOARD (Gráficos Aclarados) ───────────────────────────────────
with tab1:
    # Conteo por sector (usando datos explotados y coincidencia única de ID).
    sector_counts = exp_f.groupby("SECTOR")["ID"].nunique()

    st.markdown(
        section_title("Distribución Temática",
                       "Número único de convocatorias clasificadas por sector temático"),
        unsafe_allow_html=True,
    )

    col_a, col_b = st.columns([1.5, 1])
    with col_a:
        st.markdown(bar_chart(sector_counts, "Gráfico de Convocatorias por Sector"), unsafe_allow_html=True)
    with col_b:
        st.markdown(donut_chart(sector_counts, "Distribución de los Top 8 Sectores"), unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        if "SEGMENTO" in base_f.columns and not base_f.empty:
            st.markdown(
                donut_chart(base_f["SEGMENTO"].value_counts(), "Proporción por Segmento"),
                unsafe_allow_html=True,
            )

    if "ESTADO" in base_f.columns and not base_f.empty:
        st.markdown(section_title("Monitoreo de Estado"), unsafe_allow_html=True)
        st.markdown(
            bar_chart(base_f["ESTADO"].value_counts(), "Conteo de Registros por Estado Actual"),
            unsafe_allow_html=True,
        )


# ─── TAB 2: EXPLORADOR (Tablas Claras) ─────────────────────────────────────
with tab2:
    st.markdown(
        section_title("Base de Datos Filtrada",
                       f"Visualizando {len(base_f)} registros que cumplen los criterios actuales."),
        unsafe_allow_html=True,
    )

    # Configurar columnas a mostrar.
    id_col = "ID" if "ID" in base_f.columns else base_f.columns[0]
    show_cols = [c for c in [
        id_col, "NOMBRE DE LA CONVOCATORIA", "SEGMENTO",
        "ESTADO", "FECHA DE APERTURA", "FECHA DE CIERRE", "SECTOR",
    ] if c in base_f.columns]

    # Streamlit dataframe aplica tema por defecto (claro u oscuro).
    st.dataframe(
        base_f[show_cols].reset_index(drop=True),
        use_container_width=True,
        height=420,
        hide_index=True,
        column_config={
            "ID": st.column_config.NumberColumn("ID", width=60),
            "NOMBRE DE LA CONVOCATORIA": st.column_config.TextColumn("Nombre Convocatoria", width=350),
            "SEGMENTO": st.column_config.TextColumn("Segmento", width=180),
            "ESTADO":   st.column_config.TextColumn("Estado",   width=120),
            "SECTOR":   st.column_config.TextColumn("Sector",   width=240),
            "FECHA DE APERTURA": st.column_config.TextColumn("Apertura", width=120),
            "FECHA DE CIERRE": st.column_config.TextColumn("Cierre", width=120),
        },
    )

    st.markdown(section_title("Inspección Detallada por Sector"), unsafe_allow_html=True)
    sel_det = st.selectbox("Seleccione un sector temático", sectores_all, key="det_sector")
    if sel_det:
        # Usar dataframe explotado para este detalle.
        det = exploded_df[exploded_df["SECTOR"] == sel_det]
        
        # Volver a aplicar filtros secundarios.
        if sel_estados and "ESTADO" in det.columns:
            det = det[det["ESTADO"].isin(sel_estados)]
        if sel_segmentos and "SEGMENTO" in det.columns:
            det = det[det["SEGMENTO"].isin(sel_segmentos)]

        det_cols = [c for c in [
            id_col, "NOMBRE DE LA CONVOCATORIA", "SEGMENTO", "ESTADO",
            "FECHA DE APERTURA", "FECHA DE CIERRE", "MONTO POR PROYECTO",
        ] if c in det.columns]
        
        st.markdown(f'<div style="color:{TEXT_SUB}; font-size:0.92rem; margin-bottom:12px;">'
                    f'Se encontraron <b>{len(det)}</b> registro(s) para el sector seleccionado.'
                    f'</div>', unsafe_allow_html=True)
        
        st.dataframe(
            det[det_cols].reset_index(drop=True),
            use_container_width=True, height=280, hide_index=True,
        )


# ─── TAB 3: REPORTE EXCEL (Formatos Claros) ───────────────────────────────
with tab3:
    st.markdown(
        section_title(
            "Módulo de Exportación Automatizada",
            "Genera un documento Excel estructurado: Índice + una pestaña por cada sector temático."
        ),
        unsafe_allow_html=True,
    )

    export_mode = st.radio(
        "Datos a exportar",
        ["Exportar todos los registros originales", "Exportar únicamente registros filtrados"],
        horizontal=True,
    )
    
    export_df = exp_f if export_mode == "Exportar únicamente registros filtrados" else exploded_df

    # Vista previa de la estructura del Excel.
    preview = (
        export_df.groupby("SECTOR")["ID"]
        .nunique().reset_index()
        .rename(columns={"SECTOR": "Sector Temático", "ID": "N° Convocatorias"})
        .sort_values("Sector Temático")
    )
    
    with st.expander(
        f"Ver Estructura Planificada del Documento — {preview['Sector Temático'].nunique()} hojas planificadas"
    ):
        st.dataframe(preview, use_container_width=True, hide_index=True, height=250)

    st.markdown("<br>", unsafe_allow_html=True)

    # Usar tipo primary para botón de acción verde.
    if st.button("Construir Documento Excel", type="primary"):
        with st.spinner("Estructurando y aplicando formato al archivo Excel…"):
            excel_bytes = build_excel(export_df)
        st.success("El reporte se ha generado correctamente.")
        st.download_button(
            label="Descargar Reporte (Convocatorias_por_Sector.xlsx)",
            data=excel_bytes,
            file_name="Convocatorias_por_Sector.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
